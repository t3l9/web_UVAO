"""
Вытаскивает из public/NG/ строки, находящиеся на 8-м, 7-м или 6-м рабочем дне.

Правило (рабочие дни):
  - 8 день: дата выгрузки == регламентный срок
  - 7 день: от даты выгрузки до регламентного срока остался 1 рабочий день
  - 6 день: от даты выгрузки до регламентного срока осталось 2 рабочих дня

Дедупликация по "Номер сообщения": если одно сообщение есть в нескольких
файлах — оставляет строку из последнего (самого свежего) файла.

Запуск:
    python extract_ng_days_678.py

Опции:
    --folder  путь к папке с Excel (по умолч.: backend/public/NG/)
    --output  путь к итоговому файлу (по умолч.: рядом со скриптом)
"""

import argparse
import os
import re
from datetime import datetime, timedelta, date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FOLDER = os.path.join(SCRIPT_DIR, 'public', 'NG')

# ── Нерабочие дни (из config.py) ─────────────────────────────────────────────
EXCLUDED_DATES_STR = [
    "01.01.2026", "02.01.2026", "03.01.2026", "04.01.2026", "05.01.2026",
    "06.01.2026", "07.01.2026", "08.01.2026", "09.01.2026",
    "01.05.2026", "02.05.2026", "03.05.2026",
    "09.05.2026", "10.05.2026", "11.05.2026",
    "16.05.2026", "17.05.2026", "23.05.2026", "24.05.2026", "30.05.2026", "31.05.2026",
    "06.06.2026", "07.06.2026", "20.06.2026", "21.06.2026", "27.06.2026", "28.06.2026",
    "12.06.2026", "13.06.2026", "14.06.2026",
    "04.07.2026", "05.07.2026", "11.07.2026", "12.07.2026", "18.07.2026", "19.07.2026",
    "25.07.2026", "26.07.2026",
    "01.08.2026", "02.08.2026", "08.08.2026", "09.08.2026", "15.08.2026", "16.08.2026",
    "22.08.2026", "23.08.2026", "29.08.2026", "30.08.2026",
    "05.09.2026", "06.09.2026", "12.09.2026", "13.09.2026", "19.09.2026", "20.09.2026",
    "26.09.2026", "27.09.2026",
    "10.10.2026", "11.10.2026", "17.10.2026", "18.10.2026", "24.10.2026", "25.10.2026",
    "14.11.2026", "15.11.2026", "21.11.2026", "22.11.2026", "28.11.2026", "29.11.2026",
    "05.12.2026", "06.12.2026", "12.12.2026", "13.12.2026", "19.12.2026", "20.12.2026",
    "26.12.2026", "27.12.2026", "31.12.2026",
]
EXCLUDED = {datetime.strptime(d, "%d.%m.%Y").date() for d in EXCLUDED_DATES_STR}

# Предпочтительный порядок столбцов
PREFERRED_COLS = [
    'День',
    'Дата выгрузки',
    'Номер сообщения',
    'Номер заявки',
    'Дата публикации сообщения',
    'Регламентный срок у сообщения (Портал)',
    'Район',
    'Статус подготовки ответа на сообщение',
    'Адрес',
    'Проблемная тема',
    'Просрок (Монитор)',
    'Ответственный ОИВ первого уровня',
    'Ответственный за подготовку ответа',
]


# ── Рабочие дни ───────────────────────────────────────────────────────────────

def working_days_between(from_date: date, to_date: date) -> int:
    """
    Считает количество рабочих дней от from_date до to_date (не включая from_date,
    включая to_date). Возвращает отрицательное число если to_date < from_date.
    """
    if from_date == to_date:
        return 0
    step = 1 if to_date > from_date else -1
    count = 0
    cur = from_date + timedelta(days=step)
    while cur != to_date + timedelta(days=step):
        if cur not in EXCLUDED and cur.weekday() < 5:  # пн-пт, не праздник
            count += step
        cur += timedelta(days=step)
    return count


def classify_day(export_date: date, deadline_date: date) -> str | None:
    """
    Возвращает '8 день', '7 день', '6 день' или None.
    Рабочих дней от export_date до deadline_date (включительно):
      0 → 8 день (срок сегодня)
      1 → 7 день (завтра последний рабочий день)
      2 → 6 день (послезавтра)
    """
    wd = working_days_between(export_date, deadline_date)
    if wd == 0:
        return '8 день'
    if wd == 1:
        return '7 день'
    if wd == 2:
        return '6 день'
    return None


# ── Парсинг имени файла ───────────────────────────────────────────────────────

def parse_export_date(filename: str):
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+(\d{2}-\d{2})', filename)
    if not m:
        return None, None
    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%d.%m.%Y %H-%M")
    return dt, dt.strftime('%d.%m.%Y %H:%M')


# ── Основная логика ───────────────────────────────────────────────────────────

def run(ng_folder: str, output_path: str):
    if not os.path.isdir(ng_folder):
        print(f"Папка не найдена: {ng_folder}")
        return

    xlsx_files = [
        f for f in os.listdir(ng_folder)
        if f.lower().endswith('.xlsx') and f.startswith('Ответы в работе')
    ]

    parsed = []
    for fname in xlsx_files:
        dt, dt_str = parse_export_date(fname)
        if dt:
            parsed.append((dt, dt_str, os.path.join(ng_folder, fname)))
        else:
            print(f"  ? Пропущен (нет даты): {fname}")

    if not parsed:
        print("Нет файлов.")
        return

    parsed.sort(key=lambda x: x[0])
    print(f"Файлов: {len(parsed)}")
    print(f"Период: {parsed[0][0].strftime('%d.%m.%Y')} → {parsed[-1][0].strftime('%d.%m.%Y')}")
    print()

    DEADLINE_COL = 'Регламентный срок у сообщения (Портал)'
    frames = []

    for export_dt, export_str, filepath in parsed:
        fname = os.path.basename(filepath)
        print(f"  {fname} ...", end=' ', flush=True)

        try:
            df = pd.read_excel(filepath, sheet_name='Ответы в работе')
        except Exception as e:
            print(f"ошибка: {e}")
            continue

        if df.empty or DEADLINE_COL not in df.columns:
            print("пусто или нет столбца дедлайна")
            continue

        df[DEADLINE_COL] = pd.to_datetime(df[DEADLINE_COL], errors='coerce')
        export_d = export_dt.date()

        matched = []
        for _, row in df.iterrows():
            dl = row[DEADLINE_COL]
            if pd.isna(dl):
                continue
            day_label = classify_day(export_d, dl.date())
            if day_label:
                r = row.to_dict()
                r['День']          = day_label
                r['_export_str']   = export_str
                r['_export_dt']    = export_dt
                matched.append(r)

        print(f"{len(matched)} строк (6/7/8 день)")
        if matched:
            frames.append(pd.DataFrame(matched))

    if not frames:
        print("\nНет строк на 6-м, 7-м или 8-м дне.")
        return

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nВсего до дедупликации: {len(combined)}")

    # Дедупликация: оставить последнее вхождение по номеру сообщения
    id_col = 'Номер сообщения' if 'Номер сообщения' in combined.columns else 'Номер заявки'
    if id_col in combined.columns:
        combined.sort_values('_export_dt', ascending=True, inplace=True)
        combined.drop_duplicates(subset=[id_col], keep='last', inplace=True)
    else:
        combined.drop_duplicates(inplace=True)

    print(f"Уникальных строк:      {len(combined)}")

    # Финальный вид
    combined.drop(columns=['_export_dt'], inplace=True, errors='ignore')
    combined.rename(columns={'_export_str': 'Дата выгрузки'}, inplace=True)

    # Форматируем дату дедлайна
    if DEADLINE_COL in combined.columns:
        combined[DEADLINE_COL] = pd.to_datetime(combined[DEADLINE_COL], errors='coerce') \
                                   .dt.strftime('%d.%m.%Y')

    # Упорядочиваем столбцы
    present = [c for c in PREFERRED_COLS if c in combined.columns]
    rest    = [c for c in combined.columns if c not in PREFERRED_COLS]
    combined = combined[present + rest]

    # Сортируем: сначала 8 день, затем 7, затем 6; внутри — по дедлайну
    day_order = {'8 день': 0, '7 день': 1, '6 день': 2}
    combined['_sort_day'] = combined['День'].map(day_order).fillna(9)
    if DEADLINE_COL in combined.columns:
        combined['_sort_dl'] = pd.to_datetime(combined[DEADLINE_COL], format='%d.%m.%Y', errors='coerce')
        combined.sort_values(['_sort_day', '_sort_dl'], inplace=True)
        combined.drop(columns=['_sort_day', '_sort_dl'], inplace=True)
    else:
        combined.sort_values('_sort_day', inplace=True)
        combined.drop(columns=['_sort_day'], inplace=True)

    combined.to_excel(output_path, sheet_name='6-7-8 день', index=False)

    # ── Оформление ──────────────────────────────────────────────────────────────
    DAY_COLORS = {
        '8 день': 'FF4444',   # красный
        '7 день': 'FF9900',   # оранжевый
        '6 день': 'FFD700',   # жёлтый
    }
    DAY_TEXT = {
        '8 день': 'FFFFFF',
        '7 день': 'FFFFFF',
        '6 день': '333333',
    }

    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    thin_side   = Side(style='thin', color='CBD5E1')
    border      = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al     = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Заголовок
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_al
        cell.border = border

    # Находим индекс столбца "День"
    day_col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == 'День':
            day_col_idx = i
            break

    # Строки данных
    for row in ws.iter_rows(min_row=2):
        day_val = row[day_col_idx - 1].value if day_col_idx else None
        row_fill = PatternFill('solid', fgColor='F8F9FF')

        for cell in row:
            cell.border = border
            cell.alignment = left_al

            # Ячейку "День" красим отдельно
            if day_col_idx and cell.column == day_col_idx and day_val in DAY_COLORS:
                cell.fill = PatternFill('solid', fgColor=DAY_COLORS[day_val])
                cell.font = Font(name='Calibri', bold=True, color=DAY_TEXT[day_val], size=10)
                cell.alignment = center_al
            else:
                cell.fill = row_fill

    # Ширина столбцов
    col_widths = {
        'День':                                          10,
        'Дата выгрузки':                                 18,
        'Номер сообщения':                               20,
        'Номер заявки':                                  20,
        'Дата публикации сообщения':                     20,
        'Регламентный срок у сообщения (Портал)':        24,
        'Район':                                         18,
        'Статус подготовки ответа на сообщение':         30,
        'Адрес':                                         35,
        'Проблемная тема':                               35,
        'Просрок (Монитор)':                             20,
        'Ответственный ОИВ первого уровня':              30,
        'Ответственный за подготовку ответа':            30,
    }
    for col_idx, cell in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(str(cell.value), 20)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)

    size_kb = os.path.getsize(output_path) // 1024
    counts = combined['День'].value_counts().to_dict()
    print(f"\nГотово!")
    print(f"  8 день: {counts.get('8 день', 0)}")
    print(f"  7 день: {counts.get('7 день', 0)}")
    print(f"  6 день: {counts.get('6 день', 0)}")
    print(f"  Файл:   {output_path} ({size_kb} КБ)")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Extract 6/7/8-day NG messages')
    parser.add_argument('--folder', default=DEFAULT_FOLDER)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()

    out = args.output or os.path.join(
        SCRIPT_DIR,
        f"Срочные_6_7_8день_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    run(args.folder, out)
