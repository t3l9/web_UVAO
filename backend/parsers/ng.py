import os
import shutil
import sqlite3
import time
import win32com.client
import pythoncom
from functools import reduce
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.reader.excel import load_workbook
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from datetime import datetime, timedelta

from ..config import BASE_DIR, DATABASE_PATH, directory, login_NG, password_NG, excluded_dates, _running
from ..utils.helpers import safe_excel_operation, send_file_to_telegram, upload_reports_to_server, keep_latest_files, clean_parcing_folder
from ..utils.status import _record_success, _record_failure, _get_chromedriver


def parcing_data_lk_prefekta(attempts=2):
    for attempt in range(1, attempts + 1):
        print(f"Попытка {attempt} из {attempts}")
        driver = webdriver.Chrome(service=ChromeService(_get_chromedriver()))
        try:
            # Открываем страницу авторизации
            driver.get('https://gorod.mos.ru/api/service/auth/auth')
            username = driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]')
            password = driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]')
            username.send_keys(login_NG)
            password.send_keys(password_NG)
            login_button = driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button')
            login_button.click()

            # Ждем загрузки страницы с аналитикой
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//div[@class="dashboard__block-link"]//div[@class="button-big link"]//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]')))

            # БЛОК ДЛЯ ПРОВЕРКИ И ЗАКРЫТИЯ УВЕДОМЛЕНИЯ
            try:
                # Ждем появления кнопки уведомления (но не долго - 3 секунды)
                notification_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '/html/body/div/div/div[4]/div/div/div[2]/div[2]/div/div/button'))
                )
                # Если кнопка найдена, нажимаем на нее
                notification_button.click()
                print("Уведомление закрыто")
                time.sleep(1)  # Небольшая пауза после закрытия
            except Exception as e:
                # Если кнопка не найдена в течение 3 секунд, просто продолжаем
                print("Уведомление не обнаружено, продолжаем работу")

            # ПРОДОЛЖАЕМ ОСНОВНОЙ ПОТОК
            driver.get('https://gorod.mos.ru/admin/ker/olap/report/155')
            time.sleep(4)

            # Нажимаем кнопку для экспорта
            WebDriverWait(driver, 20).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')))
            button = driver.find_element(By.XPATH,
                                         '/html/body/div[3]/div/div[2]/div/div/div/div/form/footer/button[3]/span[2]/span')
            button.click()
            time.sleep(1)

            # Нажимаем на кнопку "Экспорт"
            button = driver.find_element(By.XPATH, "//button[contains(@class, 'bg-primary')]//span[text()='Экспорт']")
            button.click()
            time.sleep(1)

            # Переходим на страницу загрузок
            driver.get('https://gorod.mos.ru/admin/ker/olap/downloads')
            # Подождите, пока страница загрузится)
            WebDriverWait(driver, 1500).until(EC.presence_of_element_located(
                (By.XPATH,
                 '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')))
            # скачивание файла
            button = driver.find_element(By.XPATH,
                                         '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')
            button.click()
            time.sleep(2)

            print("Парсинг завершен успешно.")
            return True

        except Exception as e:
            error_msg = str(e)
            print(
                f"Произошла ошибка: {error_msg}. Повторная попытка..." if attempt < attempts else f"Попытки закончились. Ошибка: {error_msg}")
            driver.quit()
            if attempt == attempts:
                print("Парсинг не удался после всех попыток.")
                return False
        finally:
            driver.quit()


def choosing_day(excluded_date):
    today = datetime.now().date()
    user_input = today
    days_count = 8
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in
                     excluded_date]  # делаем даты удобными для прочтения, к одному формату
    # основной цикл для нахождения даты
    while days_count != 0:
        if user_input in excluded_date:
            user_input += timedelta(days=1)
        else:
            user_input += timedelta(days=1)
            days_count -= 1
    print(user_input)
    return user_input


def _sync_ng_prosrok(main_df, today_date, day_labels):
    """
    Построчно записывает текущую выгрузку "Ответы в работе" (НГ) в таблицу
    NG_prosrok: уникально по 'Номер сообщения' (дочернее сообщение), проставляет
    'День' по правилу 8 рабочих дней и помечает статус 'В работе' / 'Устранено'
    относительно предыдущей выгрузки.
    """
    day_by_date = {d: label for d, label in day_labels}
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = sqlite3.connect(DATABASE_PATH)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS NG_prosrok (
            ID TEXT PRIMARY KEY,
            PublishDate TEXT,
            District TEXT,
            Deadline TEXT,
            PreparationStatus TEXT,
            Address TEXT,
            Problem TEXT,
            MonitorOverdue TEXT,
            Day TEXT,
            Status TEXT,
            FirstSeen TEXT,
            LastSeen TEXT,
            ExportDate TEXT
        )
    """)
    try:
        cur.execute("ALTER TABLE NG_prosrok ADD COLUMN ExportDate TEXT")
    except sqlite3.OperationalError:
        pass

    def col(row, name):
        value = row.get(name) if name in main_df.columns else None
        if value is None or pd.isna(value):
            return None
        return str(value).strip() or None

    rows_to_upsert = []
    current_ids = []

    for _, row in main_df.iterrows():
        request_id = col(row, 'Номер сообщения')
        if not request_id:
            continue
        current_ids.append(request_id)

        deadline = row.get('Регламентный срок у сообщения (Портал)')
        deadline_date = deadline.date() if pd.notna(deadline) else None
        if deadline_date is not None and deadline_date < today_date:
            day_label = 'Просрок'
        else:
            day_label = day_by_date.get(deadline_date, 'Просрок')

        rows_to_upsert.append((
            request_id,
            col(row, 'Дата публикации сообщения'),
            col(row, 'Район'),
            str(deadline) if pd.notna(deadline) else None,
            col(row, 'Статус подготовки ответа на сообщение'),
            col(row, 'Адрес'),
            col(row, 'Проблемная тема'),
            col(row, 'Просрок (Монитор)'),
            day_label,
            now_str,  # FirstSeen
            now_str,  # LastSeen
            now_str,  # ExportDate — обновляется при каждой активной выгрузке
        ))

    if rows_to_upsert:
        cur.executemany("""
            INSERT INTO NG_prosrok (ID, PublishDate, District, Deadline, PreparationStatus, Address, Problem, MonitorOverdue, Day, Status, FirstSeen, LastSeen, ExportDate)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'В работе', ?, ?, ?)
            ON CONFLICT(ID) DO UPDATE SET
                PublishDate=excluded.PublishDate,
                District=excluded.District,
                Deadline=excluded.Deadline,
                PreparationStatus=excluded.PreparationStatus,
                Address=excluded.Address,
                Problem=excluded.Problem,
                MonitorOverdue=excluded.MonitorOverdue,
                Day=excluded.Day,
                Status='В работе',
                LastSeen=excluded.LastSeen,
                ExportDate=excluded.ExportDate
        """, rows_to_upsert)

    if current_ids:
        placeholders = ','.join('?' for _ in current_ids)
        cur.execute(
            f"UPDATE NG_prosrok SET Status = 'Устранено' WHERE Status != 'Устранено' AND ID NOT IN ({placeholders})",
            current_ids,
        )
    else:
        cur.execute("UPDATE NG_prosrok SET Status = 'Устранено' WHERE Status != 'Устранено'")

    conn.commit()
    conn.close()


def process_ng_prosroki_file(timenow, filepath, excluded_dates):
    # ЛОГИРОВАНИЕ
    print(f"[process_ng_prosroki_file] ЗАПУСК!")
    print(f"timenow={timenow}")
    print(f"filepath={filepath}")
    print(f"excluded_dates={excluded_dates[:3]}...")

    try:
        user_input = choosing_day(excluded_dates)
        df = pd.read_excel(filepath)
        print(f"Файл прочитан, строк: {len(df)}")
    except Exception as e:
        print(f"ОШИБКА: Не удалось прочитать Excel файл в process_ng_prosroki_file: {e}")
        print("Возможно, файл повреждён или заблокирован другим процессом.")
        raise

    df['Регламентный срок у сообщения (Портал)'] = df['Регламентный срок у сообщения (Портал)'].apply(
        lambda x: x.replace(second=0))
    df = df[df['Регламентный срок у сообщения (Портал)'] <= pd.to_datetime(user_input)]
    today = datetime.now()

    # --- Исключаем строки с "Префектура Юго-Восточного округа" из обработки ---
    mask = df['Ответственный ОИВ первого уровня'] == 'Префектура Юго-Восточного округа'
    prefect_rows = df[mask].copy()
    df_to_process = df[~mask].copy()

    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }

    districts_index = ['АВД ЮВАО', 'Выхино-Жулебино', 'Капотня', 'Кузьминки', 'Лефортово',
                       'Люблино', 'Марьино', 'Некрасовка', 'Нижегородский', 'Печатники',
                       'Рязанский', 'Текстильщики', 'Южнопортовый']

    df_to_process['Район'] = df_to_process['Ответственный ОИВ первого уровня'].map(responsible_mapping)

    valid_organizations = list(responsible_mapping.keys())
    valid_organizations.append('Префектура Юго-Восточного округа')
    df_to_process = df_to_process[df_to_process['Ответственный за подготовку ответа'].isin(valid_organizations)]

    df_final = pd.concat([prefect_rows, df_to_process])

    # просрочки ЛК ПРЕФЕКТА
    condition = (df_final['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')) & (
            df_final['Регламентный срок у сообщения (Портал)'] < today)
    prefect = df_final[condition].copy()

    pivot_prefect = pd.pivot_table(prefect, values='Номер заявки', index='Район', aggfunc='count')
    pivot_prefect = pivot_prefect.rename(columns={'Номер заявки': 'Кабинет префекта просрочки'})
    if pivot_prefect.empty:
        pivot_prefect = pd.DataFrame(index=districts_index, columns=['Кабинет префекта просрочки'])
    print(pivot_prefect)

    df_final = df_final[
        ~df_final['Ответственный за подготовку ответа'].str.contains('Префектура Юго-Восточного округа')]

    excluded_dates_with_time = [
        datetime.strptime(date_str, "%d.%m.%Y").replace(hour=23, minute=59, second=0)
        for date_str in excluded_dates
    ]
    excluded_dates_dt = pd.to_datetime(excluded_dates_with_time)
    excluded_date = [datetime.strptime(date, "%d.%m.%Y").date() for date in excluded_dates]

    main_df = df_final.copy()

    def change_status(df):
        df = df.copy()
        repl = {
            "Готовится ответ": "Готовится ответ (ОИВ взял доп. срок)",
            "На доработке": "На доработке (Город вернул)",
            "На модерации": "На модерации (Проверка города)",
            "На утверждении": "На утверждении (У куратора)",
            "Нет ответа": "Нет ответа (ОИВ не дал ответ)",
        }
        df.loc[:, "Статус подготовки ответа на сообщение"] = df[
            "Статус подготовки ответа на сообщение"].replace(repl)
        return df

    def table_is_none(date, number):
        df_empty = pd.DataFrame(index=districts_index,
                                columns=[f'{number} день ({date.strftime("%d.%m")})']).fillna(0)
        print(f"{number}-й день пустой")
        return df_empty

    def crearing_day_in_svod(df, date, number):
        new_date = date + timedelta(days=1)
        while new_date in excluded_date:
            new_date += timedelta(days=1)
        if df.empty:
            return table_is_none(new_date, number), new_date
        df_date = change_status(df[df['Регламентный срок у сообщения (Портал)'].dt.date == new_date])
        pivot_date_for_svod = pd.pivot_table(df_date, values='Номер заявки', index='Район', aggfunc='count')
        new_name = f'{number} день ({new_date.strftime("%d.%m")})'
        if not pivot_date_for_svod.empty:
            pivot_date_for_svod.rename(columns={pivot_date_for_svod.columns[-1]: new_name}, inplace=True)
            return pivot_date_for_svod, new_date
        else:
            pivot_date_for_svod = table_is_none(new_date, number)
        return pivot_date_for_svod, new_date

    # 8-й день
    today_date = datetime.now().date()
    day_8 = today_date
    while day_8 in excluded_date:
        day_8 += timedelta(days=1)
    if not main_df.empty:
        df_date_8 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_8])
    else:
        df_date_8 = pd.DataFrame(columns=main_df.columns)
    pivot8_dlya_svoda = pd.pivot_table(df_date_8, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'8 день ({day_8.strftime("%d.%m")})'
    if not pivot8_dlya_svoda.empty:
        pivot8_dlya_svoda.rename(columns={pivot8_dlya_svoda.columns[-1]: new_name}, inplace=True)
    pivot_8 = pd.pivot_table(df_date_8, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    if not pivot_8.empty:
        pivot_8.rename(columns={pivot_8.columns[-1]: 'Всего'}, inplace=True)
        pivot_8.rename(index={pivot_8.index[-1]: 'Всего'}, inplace=True)
    else:
        if pivot8_dlya_svoda.empty:
            pivot8_dlya_svoda = table_is_none(day_8, 8)

    # 7-й день
    day_7 = day_8 + timedelta(days=1)
    while day_7 in excluded_date:
        day_7 += timedelta(days=1)
    if not main_df.empty:
        df_date_7 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_7])
    else:
        df_date_7 = pd.DataFrame(columns=main_df.columns)
    pivot_7 = pd.pivot_table(df_date_7, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    pivot7_dlya_svoda = pd.pivot_table(df_date_7, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'7 день ({day_7.strftime("%d.%m")})'
    if not pivot7_dlya_svoda.empty:
        pivot7_dlya_svoda.rename(columns={pivot7_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_7.empty:
        pivot_7.rename(columns={pivot_7.columns[-1]: 'Всего'}, inplace=True)
        pivot_7.rename(index={pivot_7.index[-1]: 'Всего'}, inplace=True)
    else:
        if pivot7_dlya_svoda.empty:
            pivot7_dlya_svoda = table_is_none(day_7, 7)

    # 6-й день
    day_6 = day_7 + timedelta(days=1)
    while day_6 in excluded_date:
        day_6 += timedelta(days=1)
    if not main_df.empty:
        df_date_6 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_6])
    else:
        df_date_6 = pd.DataFrame(columns=main_df.columns)
    pivot_6 = pd.pivot_table(df_date_6, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    pivot6_dlya_svoda = pd.pivot_table(df_date_6, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'6 день ({day_6.strftime("%d.%m")})'
    if not pivot6_dlya_svoda.empty:
        pivot6_dlya_svoda.rename(columns={pivot6_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_6.empty:
        pivot_6.rename(columns={pivot_6.columns[-1]: 'Всего'}, inplace=True)
        pivot_6.rename(index={pivot_6.index[-1]: 'Всего'}, inplace=True)
    else:
        if pivot6_dlya_svoda.empty:
            pivot6_dlya_svoda = table_is_none(day_6, 6)

    # 5-й день
    day_5 = day_6 + timedelta(days=1)
    while day_5 in excluded_date:
        day_5 += timedelta(days=1)
    if not main_df.empty:
        df_date_5 = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date == day_5])
    else:
        df_date_5 = pd.DataFrame(columns=main_df.columns)
    pivot_5 = pd.pivot_table(df_date_5, values='Номер заявки', index='Район',
                             columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    pivot5_dlya_svoda = pd.pivot_table(df_date_5, values='Номер заявки', index='Район', aggfunc='count')
    new_name = f'5 день ({day_5.strftime("%d.%m")})'
    if not pivot5_dlya_svoda.empty:
        pivot5_dlya_svoda.rename(columns={pivot5_dlya_svoda.columns[-1]: new_name}, inplace=True)
    if not pivot_5.empty:
        pivot_5.rename(columns={pivot_5.columns[-1]: 'Всего'}, inplace=True)
        pivot_5.rename(index={pivot_5.index[-1]: 'Всего'}, inplace=True)
    else:
        if pivot5_dlya_svoda.empty:
            pivot5_dlya_svoda = table_is_none(day_5, 5)

    # дни 4..1
    pivot4_dlya_svoda, date4 = crearing_day_in_svod(main_df, day_5, 4)
    pivot3_dlya_svoda, date3 = crearing_day_in_svod(main_df, date4, 3)
    pivot2_dlya_svoda, date2 = crearing_day_in_svod(main_df, date3, 2)
    pivot1_dlya_svoda, date1 = crearing_day_in_svod(main_df, date2, 1)

    # ПРОСРОЧКИ
    if not main_df.empty:
        prosrok = main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date < today_date]
    else:
        prosrok = pd.DataFrame(columns=main_df.columns)
    prosrok_for_svod = pd.pivot_table(prosrok, values='Номер заявки', index='Район', aggfunc='count')
    prosrok_for_svod = prosrok_for_svod.rename(columns={'Номер заявки': 'Просрочки'})
    if prosrok_for_svod.empty:
        prosrok_for_svod = pd.DataFrame(index=districts_index, columns=['Просрочки']).fillna(0)

    df_prosrok = change_status(prosrok)
    if not df_prosrok.empty:
        pivot_prosrok = pd.pivot_table(df_prosrok, values='Номер заявки', index='Район',
                                       columns="Статус подготовки ответа на сообщение", aggfunc='count', margins=True)
    else:
        pivot_prosrok = pd.DataFrame()
    if not pivot_prosrok.empty:
        pivot_prosrok.rename(columns={pivot_prosrok.columns[-1]: 'Всего'}, inplace=True)
        pivot_prosrok.rename(index={pivot_prosrok.index[-1]: 'Всего'}, inplace=True)
    else:
        print("Просроки пустые")

    # ====== ВЫХОДНЫЕ / ПРАЗДНИЧНЫЕ ДНИ ВНУТРИ ОКНА (main_df ещё с выходными) ======
    def build_holiday_block(block_dates):
        if len(block_dates) == 1:
            label = f"Вых. дни ({block_dates[0].strftime('%d.%m')})"
        else:
            label = f"Вых. дни ({block_dates[0].strftime('%d.%m')}-{block_dates[-1].strftime('%d.%m')})"
        if not main_df.empty:
            sub = change_status(main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date.isin(block_dates)])
        else:
            sub = pd.DataFrame(columns=main_df.columns)
        piv = pd.pivot_table(sub, values='Номер заявки', index='Район', aggfunc='count')
        if not piv.empty:
            piv.rename(columns={piv.columns[-1]: label}, inplace=True)
        else:
            piv = pd.DataFrame(index=districts_index, columns=[label]).fillna(0)
        return piv, label

    day_pivots = [
        (day_8, pivot8_dlya_svoda), (day_7, pivot7_dlya_svoda),
        (day_6, pivot6_dlya_svoda), (day_5, pivot5_dlya_svoda),
        (date4, pivot4_dlya_svoda), (date3, pivot3_dlya_svoda),
        (date2, pivot2_dlya_svoda), (date1, pivot1_dlya_svoda),
    ]

    all_weekend_dates = []
    ordered_pivots = [prosrok_for_svod]
    urgent_weekend_labels = []

    # Ведущий блок: если СЕГОДНЯ выходной/праздник — дни от сегодня до 8-го дня
    leading_gap = [today_date + timedelta(days=k) for k in range((day_8 - today_date).days)]
    if leading_gap:
        all_weekend_dates.extend(leading_gap)
        block_piv, block_label = build_holiday_block(leading_gap)
        ordered_pivots.append(block_piv)
        if leading_gap[-1] < day_5:
            urgent_weekend_labels.append(block_label)

    # Блоки нерабочих дней между рабочими днями
    for i, (d, piv) in enumerate(day_pivots):
        if i > 0:
            prev_d = day_pivots[i - 1][0]
            gap = [prev_d + timedelta(days=k) for k in range(1, (d - prev_d).days)]
            if gap:
                all_weekend_dates.extend(gap)
                block_piv, block_label = build_holiday_block(gap)
                ordered_pivots.append(block_piv)
                if gap[-1] < day_5:
                    urgent_weekend_labels.append(block_label)
        ordered_pivots.append(piv)

    merged_df = reduce(lambda left, right: pd.merge(left, right, left_index=True, right_index=True, how='outer'),
                       ordered_pivots)

    # статусная разбивка по всем выходным окна (лист "Выходные статусы")
    if all_weekend_dates and not main_df.empty:
        df_weekend_all = change_status(
            main_df[main_df['Регламентный срок у сообщения (Портал)'].dt.date.isin(all_weekend_dates)])
    else:
        df_weekend_all = pd.DataFrame(columns=main_df.columns)
    pivot_weekend = pd.pivot_table(df_weekend_all, values='Номер заявки', index='Район',
                                   columns="Статус подготовки ответа на сообщение", aggfunc='count',
                                   margins=True) if not df_weekend_all.empty else pd.DataFrame()
    if not pivot_weekend.empty:
        pivot_weekend.rename(columns={pivot_weekend.columns[-1]: 'Всего'}, inplace=True)
        pivot_weekend.rename(index={pivot_weekend.index[-1]: 'Всего'}, inplace=True)

    # теперь убираем выходные из детального листа "Ответы в работе"
    holidays_df = main_df[main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)]
    main_df = main_df[~main_df['Регламентный срок у сообщения (Портал)'].isin(excluded_dates_dt)].sort_values(
        by='Регламентный срок у сообщения (Портал)')

    # ====== Синхронизация с БД для дашборда просроков НГ ======
    day_labels = [
        (day_8, '8 день'), (day_7, '7 день'), (day_6, '6 день'), (day_5, '5 день'),
        (date4, '4 день'), (date3, '3 день'), (date2, '2 день'), (date1, '1 день'),
    ]
    _sync_ng_prosrok(main_df, today_date, day_labels)

    # ====== СВОД ======
    merged_table = pd.merge(pivot_prefect, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    all_in_work = pd.DataFrame({'Всего в работе': merged_table.sum(axis=1)}).fillna(0)

    # "Всего срочных" по названиям колонок (8–5 день + просрочки + срочные выходные)
    name_8 = f'8 день ({day_8.strftime("%d.%m")})'
    name_7 = f'7 день ({day_7.strftime("%d.%m")})'
    name_6 = f'6 день ({day_6.strftime("%d.%m")})'
    name_5 = f'5 день ({day_5.strftime("%d.%m")})'
    urgent_cols = ['Кабинет префекта просрочки', 'Просрочки', name_8, name_7, name_6, name_5] + urgent_weekend_labels
    urgent_cols = [c for c in urgent_cols if c in merged_table.columns]
    all_urgent = pd.DataFrame({'Всего срочных': merged_table[urgent_cols].sum(axis=1)}).fillna(0)

    final_svod = pd.merge(all_in_work, pivot_prefect, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, all_urgent, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = pd.merge(final_svod, merged_df, left_index=True, right_index=True, how='outer').fillna(0)
    final_svod = final_svod.sort_values(by='Всего срочных', ascending=False)

    totals_row = final_svod.sum(axis=0)
    totals_row.name = 'Итог по округу'
    df_with_totals = pd.concat([final_svod, pd.DataFrame(totals_row).T])
    df_with_totals.index.name = 'Ответственный за подготовку ответа'

    # сохраняем по пути и добавляем листы
    processed_file_path = os.path.join(directory, f"Ответы в работе {timenow}.xlsx")
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        df_with_totals.to_excel(writer, sheet_name='СВОД', index=True, startrow=2)
        pivot_prosrok.to_excel(writer, sheet_name='просрочки', index=True, startrow=2)
        pivot_8.to_excel(writer, sheet_name='8-й день', index=True, startrow=2)
        pivot_7.to_excel(writer, sheet_name='7-й день', index=True, startrow=2)
        pivot_6.to_excel(writer, sheet_name='6-й день', index=True, startrow=2)
        pivot_5.to_excel(writer, sheet_name='5-й день', index=True, startrow=2)
        pivot_weekend.to_excel(writer, sheet_name='Выходные статусы', index=True, startrow=2)  # НОВЫЙ ЛИСТ
        main_df.to_excel(writer, sheet_name='Ответы в работе', index=False, startrow=0)
        holidays_df.to_excel(writer, sheet_name='Выходные', index=False, startrow=0)
        prefect.to_excel(writer, sheet_name='Префект просрок', index=False, startrow=0)


def personalizating_table_osn(timestamp):
    from openpyxl.utils import get_column_letter

    timenow = datetime.now().strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")
    wb = load_workbook(file_path)
    ws = wb.worksheets[0]

    header_row = 3
    data_first = 4
    total_row = ws.max_row  # строка "Итог по округу"
    data_last = total_row - 1
    last_col = ws.max_column
    last_col_letter = get_column_letter(last_col)

    light_blue_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    pale_blue_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    pink_fill = PatternFill(start_color="f7867e", end_color="f7867e", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def tnr(bold=False, color=None):
        return Font(name='Times New Roman', size=11, bold=bold, color=color)

    # заголовок отчёта (строка 2)
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws['A2'] = (f'Сводная информация по нарушениям сроков подготовки ответов на сообщения, поступившие на '
                f'централизованный портал "Наш город" по состоянию на {timenow} '
                f'{datetime.now().strftime("%d.%m.%y")} г.')
    ws['A2'].font = tnr(bold=True)
    ws['A2'].alignment = center
    for cell in ws[f'A2:{last_col_letter}2'][0]:
        cell.border = thin_border
    ws.row_dimensions[2].height = 37

    # шапка таблицы (строка 3)
    for cell in ws[header_row]:
        cell.fill = light_blue_fill
        cell.font = tnr(bold=True)
        cell.border = thin_border
        cell.alignment = center
    ws.row_dimensions[header_row].height = 55

    # красные / голубые столбцы по заголовкам
    titles = {c.column: str(c.value or "") for c in ws[header_row]}
    idx_5 = next((i for i, t in titles.items() if t.startswith('5 день')), None)
    pink_cols, pale_extra_cols = [], []
    for i, t in titles.items():
        if i == 1:
            continue
        if t in ('Кабинет префекта просрочки', 'Просрочки') or t.startswith(('8 день', '7 день', '6 день', '5 день')):
            pink_cols.append(i)
        elif t.startswith('Вых. дни'):
            if idx_5 is not None and i < idx_5:  # выходные в диапазоне 8–5 дня
                pink_cols.append(i)
            else:
                pale_extra_cols.append(i)
        elif t.startswith(('4 день', '3 день', '2 день', '1 день')):
            pale_extra_cols.append(i)

    # данные
    for row in ws.iter_rows(min_row=data_first, max_row=data_last, min_col=1, max_col=last_col):
        for cell in row:
            cell.font = tnr()
            cell.border = thin_border
            cell.alignment = center
            col = cell.column
            if col == 1:  # район
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col == 2:  # Всего в работе
                cell.fill = pale_blue_fill
            elif col == 3:  # Кабинет префекта просрочки
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True)
            elif col == 4:  # Всего срочных
                cell.fill = pale_blue_fill
                cell.font = tnr(bold=True, color="800000")
            elif col == 5:  # Просрочки
                cell.font = tnr(bold=True)
            elif col in pale_extra_cols:  # дни 4–1 и обычные выходные
                cell.fill = pale_blue_fill

    # итоговая строка
    for cell in ws[total_row]:
        cell.font = tnr(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # условная заливка красным (>0)
    for col_idx in pink_cols:
        cl = get_column_letter(col_idx)
        rng = f"{cl}{data_first}:{cl}{data_last}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=pink_fill))

    # ширина столбцов
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    for col in range(6, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 7.5

    # высота строк
    for r in range(data_first, total_row + 1):
        ws.row_dimensions[r].height = 14.5

    wb.save(file_path)
    print(f'Formatting applied to the first table in {file_path} successfully.')


def personalizating_table_weekend(timestamp):
    """Лист "Выходные статусы" (в PDF: "Сообщения в выходные дни в разрезе по статусам").
    worksheets[6] (7-й по счёту). ВЫЗЫВАТЬ ПЕРЕД add_run_delete_and_save_files."""
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")
    wb = load_workbook(file_path)
    ws = wb.worksheets[6]  # 0:СВОД 1:просрочки 2:8д 3:7д 4:6д 5:5д 6:ВЫХОДНЫЕ СТАТУСЫ

    start_row = 3
    max_row = ws.max_row
    max_column = ws.max_column

    header_fill = PatternFill(start_color="5286d1", end_color="5286d1", fill_type="solid")
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc", fill_type="solid")
    red_font = Font(color="FF0000", bold=True)
    bold_font = Font(bold=True)
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    if max_column > 0:
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Сообщения в выходные дни в разрезе по статусам'
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    header_row = ws[start_row]
    for cell in header_row:
        cell.fill = header_fill
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in row:
            cell.fill = body_fill
            if cell.column != 1 and cell.column != max_column:
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue
                if row[0].row == max_row:
                    continue
                cell.font = red_font
    wb.save(file_path)
    print(f'Formatting applied to the weekend-status sheet in {file_path} successfully.')


def personalizating_table_prosrok(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[1]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = 'Просроченные сообщения в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_eight_day(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[2]  # Индексация начинается с 0, поэтому 1 - это второй лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '8-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_seven_day(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[3]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '7-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_six_day(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[4]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '6-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the first table in the first sheet in {file_path} successfully.')


def personalizating_table_five_day(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")

    # Открытие существующего файла
    wb = load_workbook(file_path)
    ws = wb.worksheets[5]  # Индексация начинается с 0, поэтому 3 - это 4 лист

    # Определяем начальную ячейку таблицы
    start_row = 3

    # Определяем ширину таблицы (количество столбцов)
    max_row = ws.max_row
    max_column = ws.max_column

    # Определяем стили
    header_fill = PatternFill(start_color="5286d1", end_color="5286d1",
                              fill_type="solid")  # Синеватый цвет для заголовка
    body_fill = PatternFill(start_color="bdd7fc", end_color="bdd7fc",
                            fill_type="solid")  # Голубой цвет для остальных строк
    red_font = Font(color="FF0000", bold=True)  # Красный цвет для текста
    bold_font = Font(bold=True)  # Жирный шрифт
    border = Border(left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"))

    # Добавляем название таблицы
    # Проверяем, есть ли данные и больше ли 0 столбцов
    if max_column > 0:
        # Объединяем ячейки для заголовка (одна строка выше)
        header_range = f"A{start_row - 1}:{chr(64 + max_column)}{start_row - 1}"  # Объединяем ячейки в строке выше
        ws.merge_cells(header_range)
        ws[f'A{start_row - 1}'] = '5-й день в разрезе по статусам'

        # Применяем форматирование к заголовку
        header_cell = ws[f'A{start_row - 1}']
        header_cell.font = Font(name='Times New Roman', bold=True, size=11)
        header_cell.alignment = Alignment(horizontal='center', vertical='center',
                                          wrap_text=True)  # Устанавливаем выравнивание и перенос текста

        # Добавляем черные границы ко всему диапазону заголовка
        thin = Side(border_style="thin", color="000000")  # Черная граница
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws[header_range.split(':')[0]:header_range.split(':')[1]]:
            for cell in row:
                cell.border = border

    # Применяем стили к заголовку таблицы (строка start_row)
    header_row = ws[start_row]  # Заголовок таблицы находится в строке start_row
    for cell in header_row:
        cell.fill = header_fill  # Синеватая заливка
        cell.font = Font(name='Times New Roman', bold=True, size=9)
        cell.border = border  # Границы для всех ячеек заголовка
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)  # Центрируем текст и включаем перенос
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    # Применяем стили к строкам таблицы
    for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            cell.font = bold_font  # Жирный текст
            cell.border = border  # Черные границы
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)  # Центрируем текст и включаем перенос

        # Устанавливаем цвет фона и текста
        for cell in row:
            cell.fill = body_fill  # Голубой фон для остальных строк
            # Проверяем, является ли столбец "На модерации" или "На утверждении"
            if cell.column != 1 and cell.column != max_column:  # Исключаем первый и последний столбцы
                if header_row[cell.column - 1].value in ["На модерации (Проверка города)",
                                                         "На утверждении (У куратора)"]:
                    continue  # Пропускаем эти столбцы
                # Исключаем последнюю строку
                if row[0].row == max_row:
                    continue  # Пропускаем последнюю строку
                cell.font = red_font  # Красный текст
    wb.save(file_path)
    print(f'Formatting applied to the five table in the first sheet in {file_path} successfully.')


def add_run_delete_and_save_files(timestamp):
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    file_path = os.path.join(directory, f"Ответы в работе {timestamp}.xlsx")
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(os.path.abspath(file_path))

    # Макрос: добавлен лист 7 (Выходные статусы) — копируется в СВОД и удаляется
    vba_code = """
    Sub CopyTablesToFirstSheet()
                Dim wsFirst As Worksheet
                Dim wsSecond As Worksheet
                Dim wsThird As Worksheet
                Dim wsFour As Worksheet
                Dim wsFive As Worksheet
                Dim wsSix As Worksheet
                Dim wsSeven As Worksheet
                Dim lastRow As Long
                Dim copyRange As Range
                Set wsFirst = ThisWorkbook.Worksheets(1)  ' СВОД
                Set wsSecond = ThisWorkbook.Worksheets(2) ' просрочки
                Set wsThird = ThisWorkbook.Worksheets(3)  ' 8-й день
                Set wsFour = ThisWorkbook.Worksheets(4)   ' 7-й день
                Set wsFive = ThisWorkbook.Worksheets(5)   ' 6-й день
                Set wsSix = ThisWorkbook.Worksheets(6)    ' 5-й день
                Set wsSeven = ThisWorkbook.Worksheets(7)  ' Выходные статусы
                ' просрочки
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsSecond.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                ' 8-й день
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsThird.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                ' 7-й день
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsFour.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                ' 6-й день
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsFive.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                ' 5-й день
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsSix.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                ' Выходные статусы
                lastRow = wsFirst.Cells(wsFirst.Rows.Count, 1).End(xlUp).Row + 2
                Set copyRange = wsSeven.UsedRange
                copyRange.Copy wsFirst.Cells(lastRow, 1)
                Application.CutCopyMode = False
                Application.DisplayAlerts = False
                wsSecond.Delete
                wsThird.Delete
                wsFour.Delete
                wsFive.Delete
                wsSix.Delete
                wsSeven.Delete
                Application.DisplayAlerts = True
            End Sub

            Sub DeleteFirstSheet()
                Dim wsFirst As Worksheet
                Set wsFirst = ThisWorkbook.Worksheets(1)
                Application.DisplayAlerts = False
                wsFirst.Delete
                Application.DisplayAlerts = True
            End Sub
        """

    vba_module = wb.VBProject.VBComponents.Add(1)
    vba_module.Name = 'MyMacroModule'
    vba_module.CodeModule.AddFromString(vba_code)

    excel.Application.Run('MyMacroModule.CopyTablesToFirstSheet')

    # PDF
    pdf_file_name = f"Ответы в работе {timestamp}.pdf"
    pdf_path = os.path.join(os.path.dirname(file_path), pdf_file_name)
    wsFirst = wb.Worksheets(1)

    wsFirst.PageSetup.FitToPagesWide = 1
    wsFirst.PageSetup.FitToPagesTall = 1
    wsFirst.PageSetup.Zoom = False
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    wb.Save()
    try:
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)
            print("Файл успешно удален.")
        print(f"Сохранение файла в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении файла: {e}")

    excel.Application.Run('MyMacroModule.DeleteFirstSheet')

    for sheet in wb.Worksheets:
        sheet.Cells.EntireColumn.AutoFit()

    wb.Save()
    wb.Close()
    excel.Quit()

    # ЛОГИРОВАНИЕ ПЕРЕД ПЕРЕМЕЩЕНИЕМ
    print(f"[add_run_delete_and_save_files] ПЕРЕД ПЕРЕМЕЩЕНИЕМ")
    print(f"file_path={file_path}")
    print(f"pdf_path={pdf_path}")
    print(f"file_path exists={os.path.exists(file_path)}")
    print(f"pdf_path exists={os.path.exists(pdf_path)}")

    static_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'public/NG')
    print(f"[DEBUG NG] static_directory = {static_directory}")

    if not os.path.exists(static_directory):
        os.makedirs(static_directory)
        print(f"Папка создана: {static_directory}")
    else:
        print(f"Папка существует: {static_directory}")

    if os.path.exists(os.path.join(static_directory, os.path.basename(file_path))):
        os.remove(os.path.join(static_directory, os.path.basename(file_path)))
    shutil.move(file_path, static_directory)
    print(f"Файл Excel перемещен в: {static_directory} ({os.path.basename(file_path)})")

    if os.path.exists(os.path.join(static_directory, os.path.basename(pdf_path))):
        os.remove(os.path.join(static_directory, os.path.basename(pdf_path)))
    shutil.move(pdf_path, static_directory)
    print(f"Файл PDF перемещен в: {static_directory} ({os.path.basename(pdf_path)})")

    files_in_folder = os.listdir(static_directory)
    print(f"Файлов в папке NG: {files_in_folder}")

    return pdf_path, file_path


# ──────────────────────────────────────────────
# VBA-макрос (две сводные + оформление)
# ──────────────────────────────────────────────

VBA_MACRO = r"""
Sub CreateReport()
    Dim wsData As Worksheet, wsPivot As Worksheet
    Dim pc As PivotCache
    Dim pt1 As PivotTable, pt2 As PivotTable
    Dim lastRow As Long, lastCol As Long, startRow2 As Long

    Set wsData = ThisWorkbook.Sheets("Sheet1")

    ' Формат дат в исходных данных
    wsData.Columns("B").NumberFormat = "DD.MM.YYYY"  ' Дата отображения (Монитор)
    wsData.Columns("C").NumberFormat = "DD.MM.YYYY"  ' Регламентный срок (Портал)

    ' Пересоздаём лист сводных
    Application.DisplayAlerts = False
    On Error Resume Next
    ThisWorkbook.Sheets("Сводная таблица").Delete
    On Error GoTo 0
    Application.DisplayAlerts = True
    Set wsPivot = ThisWorkbook.Sheets.Add
    wsPivot.Name = "Сводная таблица"

    lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row
    lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column

    ' Один кэш на обе сводные
    Dim srcAddr As String
    srcAddr = wsData.Name & "!" & wsData.Cells(1, 1).Resize(lastRow, lastCol).Address
    Set pc = ThisWorkbook.PivotCaches.Create( _
        SourceType:=xlDatabase, _
        SourceData:=srcAddr)

    ' Заголовок отчёта
    With wsPivot.Range("A1")
        .Value = "Отчёт ЛК Префекта ЮВАО"
        .Font.Name = "Times New Roman": .Font.Size = 14: .Font.Bold = True
    End With
    With wsPivot.Range("A2")
        .Value = "Сформировано: " & Format(Now, "DD.MM.YYYY HH:MM")
        .Font.Name = "Times New Roman": .Font.Size = 9: .Font.Italic = True
    End With
    With wsPivot.Range("A3")
        .Value = "Срок ответа по районам (Регламентный срок, Портал)"
        .Font.Name = "Times New Roman": .Font.Bold = True
    End With

    ' ===== Сводная 1: Регламентный срок x Район =====
    Set pt1 = pc.CreatePivotTable(TableDestination:=wsPivot.Cells(4, 1), TableName:="PT_Срок")
    With pt1
        .PivotFields("Район").Orientation = xlColumnField
        .PivotFields("Регламентный срок у сообщения (Портал)").Orientation = xlRowField
        .AddDataField .PivotFields("Номер заявки"), "Кол-во", xlCount
        .RowAxisLayout xlTabularRow
        .ColumnGrand = True: .RowGrand = True
    End With
    On Error Resume Next
    pt1.PivotFields("Регламентный срок у сообщения (Портал)").DataRange.Cells(1).Ungroup
    ' Формат дат задаём на самих ячейках-метках (PivotField.NumberFormat у строкового поля даёт 1004)
    pt1.PivotFields("Регламентный срок у сообщения (Портал)").DataRange.NumberFormat = "DD.MM.YYYY"
    On Error GoTo 0
    StylePivot pt1

    ' ===== Сводная 2: Дата отображения x Район =====
    startRow2 = pt1.TableRange2.Row + pt1.TableRange2.Rows.Count + 2
    With wsPivot.Cells(startRow2, 1)
        .Value = "Количество сообщений по дате отображения (Монитор)"
        .Font.Name = "Times New Roman": .Font.Size = 12: .Font.Bold = True
    End With

    Set pt2 = pc.CreatePivotTable(TableDestination:=wsPivot.Cells(startRow2 + 1, 1), TableName:="PT_Дата")
    With pt2
        .PivotFields("Район").Orientation = xlColumnField
        .PivotFields("Дата отображения (Монитор)").Orientation = xlRowField
        .AddDataField .PivotFields("Номер заявки"), "Кол-во", xlCount
        .RowAxisLayout xlTabularRow
        .ColumnGrand = True: .RowGrand = True
    End With
    ' Разгруппировать даты -> вид ДД.ММ.ГГГГ
    On Error Resume Next
    pt2.PivotFields("Дата отображения (Монитор)").DataRange.Cells(1).Ungroup
    On Error GoTo 0
    ' Скрыть пустые значения "(пусто)" в таблице с монитором
    HideBlankItems pt2, "Дата отображения (Монитор)"
    HideBlankItems pt2, "Район"
    On Error Resume Next
    ' Формат дат задаём на самих ячейках-метках (PivotField.NumberFormat у строкового поля даёт 1004)
    pt2.PivotFields("Дата отображения (Монитор)").DataRange.NumberFormat = "DD.MM.YYYY"
    ' Хронология: от ранней даты к поздней (по возрастанию)
    pt2.PivotFields("Дата отображения (Монитор)").AutoSort xlAscending, "Дата отображения (Монитор)"
    On Error GoTo 0
    StylePivot pt2

    ' Столбец A держит подписи полей строк ("Регламентный срок у сообщения (Портал)",
    ' "Дата отображения (Монитор)") — делаем его широким, чтобы названия помещались.
    wsPivot.Columns("A").ColumnWidth = 32

    With wsPivot.PageSetup
        .Orientation = xlLandscape
        .FitToPagesWide = 1: .FitToPagesTall = False
        .LeftMargin = Application.CentimetersToPoints(0.5)
        .RightMargin = Application.CentimetersToPoints(0.5)
        .TopMargin = Application.CentimetersToPoints(0.5)
        .BottomMargin = Application.CentimetersToPoints(0.5)
    End With
End Sub

Private Sub StylePivot(pt As PivotTable)
    pt.TableStyle2 = "PivotStyleLight16"
    With pt.TableRange2
        .Font.Name = "Times New Roman": .Font.Size = 10
        .HorizontalAlignment = xlCenter: .VerticalAlignment = xlCenter
        .WrapText = True
        ' Границы у всех ячеек таблицы (внешние + внутренние)
        With .Borders
            .LineStyle = xlContinuous
            .Weight = xlThin
            .Color = RGB(140, 140, 140)
        End With
    End With

    Dim body As Range, heat As Range
    On Error Resume Next
    Set body = pt.DataBodyRange
    On Error GoTo 0
    If Not body Is Nothing Then
        body.FormatConditions.Delete

        ' Исключаем строку и столбец общего итога из заливки.
        ' DataBodyRange включает итоги: последняя строка = "Общий итог" (строки),
        ' последний столбец = "Общий итог" (столбцы). Обрезаем их.
        Set heat = body
        If pt.ColumnGrand And heat.Rows.Count > 1 Then
            Set heat = heat.Resize(heat.Rows.Count - 1, heat.Columns.Count)
        End If
        If pt.RowGrand And heat.Columns.Count > 1 Then
            Set heat = heat.Resize(heat.Rows.Count, heat.Columns.Count - 1)
        End If

        ' Адаптивный градиент: минимум -> жёлтый, максимум -> красный.
        ' LowestValue/HighestValue пересчитываются от реальных чисел (без итогов),
        ' поэтому диапазон 1..10 и 1..70 красятся одинаково корректно.
        heat.FormatConditions.AddColorScale ColorScaleType:=2
        With heat.FormatConditions(heat.FormatConditions.Count)
            .ColorScaleCriteria(1).Type = xlConditionValueLowestValue
            .ColorScaleCriteria(1).FormatColor.Color = RGB(255, 243, 178)  ' пастельно-жёлтый (минимум)
            .ColorScaleCriteria(2).Type = xlConditionValueHighestValue
            .ColorScaleCriteria(2).FormatColor.Color = RGB(244, 169, 160)  ' пастельно-красный (максимум)
        End With
    End If
    ' Автоширину применяем к столбцам данных, но НЕ к A —
    ' столбец A держит длинные подписи полей и задаётся вручную.
    On Error Resume Next
    Intersect(pt.TableRange2, pt.TableRange2.Offset(0, 1)).EntireColumn.AutoFit
    On Error GoTo 0
End Sub

Private Sub HideBlankItems(pt As PivotTable, fieldName As String)
    Dim pi As PivotItem
    On Error Resume Next
    For Each pi In pt.PivotFields(fieldName).PivotItems
        If Trim(pi.Caption) = "" Or pi.Caption = "(пусто)" Or LCase(pi.Caption) = "(blank)" Then
            pi.Visible = False
        End If
    Next pi
    On Error GoTo 0
End Sub
"""


def process_lk_prefekta_file(filepath, directory, timestamp, selected_district="Все районы"):
    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        print(f"ОШИБКА: Не удалось прочитать Excel файл в process_lk_prefekta_file: {e}")
        print("Возможно, файл повреждён или заблокирован другим процессом.")
        raise  # Пробрасываем ошибку выше

    responsible_mapping = {
        'ГБУ «Автомобильные дороги ЮВАО»': 'АВД ЮВАО',
        'ГБУ Жилищник Выхино района Выхино-Жулебино города Москвы': 'Выхино-Жулебино',
        'Управа Выхино-Жулебино': 'Выхино-Жулебино',
        'ГБУ Жилищник Нижегородского района города Москвы': 'Нижегородский',
        'Управа Нижегородский': 'Нижегородский',
        'ГБУ Жилищник района Капотня города Москвы': 'Капотня',
        'Управа Капотня': 'Капотня',
        'ГБУ Жилищник района Кузьминки города Москвы': 'Кузьминки',
        'Управа Кузьминки': 'Кузьминки',
        'ГБУ Жилищник района Лефортово города Москвы': 'Лефортово',
        'Управа Лефортово': 'Лефортово',
        'ГБУ Жилищник района Люблино города Москвы': 'Люблино',
        'Управа Люблино': 'Люблино',
        'ГБУ Жилищник района Марьино города Москвы': 'Марьино',
        'Управа Марьино': 'Марьино',
        'ГБУ Жилищник района Некрасовка города Москвы': 'Некрасовка',
        'Управа Некрасовка': 'Некрасовка',
        'ГБУ Жилищник района Печатники города Москвы': 'Печатники',
        'Управа Печатники': 'Печатники',
        'ГБУ Жилищник района Текстильщики города Москвы': 'Текстильщики',
        'Управа Текстильщики': 'Текстильщики',
        'ГБУ Жилищник Рязанского района города Москвы': 'Рязанский',
        'Управа Рязанский': 'Рязанский',
        'ГБУ Жилищник Южнопортового района города Москвы': 'Южнопортовый',
        'Управа Южнопортовый': 'Южнопортовый'
    }

    # Функция для обновления значений в столбце 'Район'
    def update_region(row):
        if row['Ответственный ОИВ первого уровня'] == 'Префектура Юго-Восточного округа':
            return row['Район']  # Ничего не меняем
        else:
            return responsible_mapping.get(row['Ответственный ОИВ первого уровня'], row['Район'])

    # Применение функции к каждому ряду
    df['Район'] = df.apply(update_region, axis=1)

    df_filtered = df[df['Ответственный за подготовку ответа'] == 'Префектура Юго-Восточного округа']

    columns_to_keep = [
        "Номер заявки",
        "Номер сообщения",
        "Дата отображения (Монитор)",
        "Регламентный срок у сообщения (Портал)",
        "Признак Монитора",
        "Просрок (Монитор)",
        "Дата публикации сообщения",
        "Район",
        "Проблемная тема",
        "Адрес",
        "Категория объекта",
        "Категория/действие последнего ответа",
        "Ответственный за подготовку ответа",
        "Ответственный ОИВ первого уровня",
        "Статус подготовки ответа на сообщение"
    ]

    # Проверяем, какие столбцы из списка реально существуют
    existing_columns = [col for col in columns_to_keep if col in df_filtered.columns]
    missing_columns = [col for col in columns_to_keep if col not in df_filtered.columns]

    if missing_columns:
        print(f"Внимание! Отсутствуют следующие столбцы: {missing_columns}")
        print("Доступные столбцы:", df_filtered.columns.tolist())

    # Используем только существующие столбцы
    df_filtered = df_filtered[existing_columns]

    if selected_district != "Все районы" and 'Район' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Район'] == selected_district]

    df_filtered = df_filtered.dropna(how='all')
    if df_filtered.empty:
        print("После фильтрации не осталось данных.")
        return None


    now = pd.Timestamp.now()
    today = datetime.now()
    timenow = today.strftime("%H-%M")
    processed_file_path = os.path.join(directory, f"ЛК Префекта {timestamp}.xlsx")
    print(f"Saving processed file to: {processed_file_path}")
    df_filtered.to_excel(processed_file_path, index=False)
    excel_file = processed_file_path


    # Запускаем Excel
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False

    workbook = excel.Workbooks.Open(excel_file)

    # Добавляем новый модуль VBA и вставляем макрос
    vb_module = workbook.VBProject.VBComponents.Add(1)  # 1 = стандартный модуль
    vb_module.CodeModule.AddFromString(VBA_MACRO)

    # Выполняем макрос
    excel.Application.Run("CreateReport")
    print("Сводные таблицы созданы")

    # Создание PDF
    pdf_file_name = f"ЛК Префекта {timestamp}.pdf"
    pdf_path = os.path.join(os.path.dirname(processed_file_path), pdf_file_name)

    wsFirst = workbook.Worksheets(1)  # Ссылка на первый лист

    # Настройки страницы для печати
    wsFirst.PageSetup.FitToPagesWide = 1
    wsFirst.PageSetup.FitToPagesTall = 1
    wsFirst.PageSetup.Zoom = False

    # Уменьшаем поля
    wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
    wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)
    workbook.Save()

    try:
        if os.path.exists(pdf_path):
            print(f"Файл {pdf_path} существует. Удаление...")
            os.remove(pdf_path)
            print("Файл успешно удалён.")

        print(f"Сохранение PDF в {pdf_path}...")
        wsFirst.ExportAsFixedFormat(0, pdf_path)  # 0 = xlTypePDF
        print(f"PDF успешно создан: {pdf_path}")
    except Exception as e:
        print(f"Ошибка при сохранении PDF: {e}")

    # Автоширина для листа со сводными
    try:
        sheet = workbook.Worksheets(2)
        sheet.Cells.EntireColumn.AutoFit()
    except Exception as e:
        print(f"Не удалось применить автоширину: {e}")

    # Сохраняем и закрываем
    workbook.Save()
    workbook.Close()
    excel.Quit()

    # Перемещение файлов в папку static
    static_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'public/Pref')

    # ЛОГИРОВАНИЕ для отладки
    print(f"[DEBUG Pref] __file__ = {__file__}")
    print(f"[DEBUG Pref] dirname = {os.path.dirname(os.path.abspath(__file__))}")
    print(f"[DEBUG Pref] static_directory = {static_directory}")

    if not os.path.exists(static_directory):
        os.makedirs(static_directory)
        print(f"Папка создана: {static_directory}")
    else:
        print(f"Папка существует: {static_directory}")

    # Перемещение Excel файла
    if os.path.exists(os.path.join(static_directory, os.path.basename(processed_file_path))):
        os.remove(os.path.join(static_directory, os.path.basename(processed_file_path)))
    shutil.move(processed_file_path, static_directory)
    print(f"Файл Excel перемещен в: {static_directory}")
    print(f"   Имя файла: {os.path.basename(processed_file_path)}")

    # Перемещение PDF файла
    if os.path.exists(os.path.join(static_directory, os.path.basename(pdf_path))):
        os.remove(os.path.join(static_directory, os.path.basename(pdf_path)))
    shutil.move(pdf_path, static_directory)
    print(f"Файл PDF перемещен в: {static_directory}")
    print(f"   Имя файла: {os.path.basename(pdf_path)}")

    # Проверка: список файлов в папке
    files_in_folder = os.listdir(static_directory)
    print(f"Файлов в папке Pref: {files_in_folder}")

    return processed_file_path


def ng(scheduled_time=None):
    """
    Обработка отчётов: Наш Город + ЛК Префекта
    """
    if _running['ng']:
        print("[ng] Пропуск: предыдущий запуск ещё выполняется")
        return
    _running['ng'] = True
    _coinit = False
    try:
        # Время, в которое нужно отправлять отчёты
        special_times_pref = {"08:55", "15:55"}
        is_special = scheduled_time in special_times_pref
        print(f"Запуск ng() в {scheduled_time}, спец запуск? → {is_special}")

        # Фиксируем временную метку в начале выполнения
        timestamp = datetime.now().strftime("%d.%m.%Y %H-%M")

        # 1. Парсинг данных
        if not parcing_data_lk_prefekta():
            _record_failure('ng', 'Парсер не смог загрузить данные после всех попыток')
            return
        pythoncom.CoInitialize()
        _coinit = True
        files = os.listdir(directory)
        files.sort(key=lambda x: os.path.getmtime(os.path.join(directory, x)))
        latest_downloaded_file = files[-1]
        source_path = os.path.join(directory, latest_downloaded_file)

        # Перемещение в Desktop/parcing
        desktop_folder = os.path.join(os.path.expanduser("~"), "Desktop", "parcing")
        os.makedirs(desktop_folder, exist_ok=True)
        dest_path = os.path.join(desktop_folder, latest_downloaded_file)
        shutil.move(source_path, dest_path)
        filepath = dest_path

        # 2. Обработка "Ответы в работе" (NG) - создаёт Excel + PDF + перемещает в public/NG/
        print(f"[ng()] Обработка Ответы в работе...")
        process_ng_prosroki_file(timestamp, filepath, excluded_dates)
        print(f"process_ng_prosroki_file завершена")

        # 3. Персонализация таблиц
        print(f"[ng()] Персонализация таблиц...")
        personalizating_table_osn(timestamp)
        personalizating_table_prosrok(timestamp)
        personalizating_table_eight_day(timestamp)
        personalizating_table_seven_day(timestamp)
        personalizating_table_six_day(timestamp)
        personalizating_table_five_day(timestamp)
        personalizating_table_weekend(timestamp)
        add_run_delete_and_save_files(timestamp)
        print(f"Персонализация завершена")

        # 4. Обработка "ЛК Префекта" (Pref) - создаёт Excel + PDF + перемещает в public/Pref/
        print(f"[ng()] Обработка ЛК Префекта...")
        process_lk_prefekta_file(filepath, directory, timestamp)
        print(f"process_lk_prefekta_file завершена")

        # 5. Отправка в Telegram (только в спец. время)
        if is_special:
            print("→ Специальное время — отправляем отчёты")

            base_dir = os.path.dirname(os.path.abspath(__file__))
            pref_dir = os.path.join(base_dir, "..", "public", "Pref")
            ng_dir = os.path.join(base_dir, "..", "public", "NG")

            # Получаем последний Excel из Pref
            pref_files = os.listdir(pref_dir)
            pref_excel_files = [f for f in pref_files if f.lower().endswith(".xlsx")]
            pref_excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(pref_dir, x)))
            pref_excel_path = os.path.join(pref_dir, pref_excel_files[-1])

            send_file_to_telegram(pref_excel_path, caption="ЛК Префекта")

            # Получаем последний Excel и PDF из NG
            ng_files = os.listdir(ng_dir)
            ng_excel_files = [f for f in ng_files if f.lower().endswith(".xlsx")]
            ng_pdf_files = [f for f in ng_files if f.lower().endswith(".pdf")]
            ng_excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(ng_dir, x)))
            ng_pdf_files.sort(key=lambda x: os.path.getmtime(os.path.join(ng_dir, x)))
            ng_excel_path = os.path.join(ng_dir, ng_excel_files[-1])
            ng_pdf_path = os.path.join(ng_dir, ng_pdf_files[-1])

            send_file_to_telegram(ng_excel_path, caption="Ответы в работе — Excel")
            send_file_to_telegram(ng_pdf_path, caption="Ответы в работе — PDF")

            print("Все файлы отправлены в Telegram")

        # 6. Загрузка отчётов на сервер (ТОЛЬКО файлы, БЕЗ сборки React!)
        print(f"[ng()] Загрузка отчётов на сервер...")
        base_dir = os.path.dirname(os.path.abspath(__file__))
        pref_dir = os.path.join(base_dir, "..", "public", "Pref")
        ng_dir = os.path.join(base_dir, "..", "public", "NG")

        # Получаем файлы из Pref
        pref_files = os.listdir(pref_dir)
        pref_excel_files = [f for f in pref_files if f.lower().endswith(".xlsx")]
        pref_pdf_files = [f for f in pref_files if f.lower().endswith(".pdf")]
        pref_excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(pref_dir, x)))
        pref_pdf_files.sort(key=lambda x: os.path.getmtime(os.path.join(pref_dir, x)))
        pref_excel_path = os.path.join(pref_dir, pref_excel_files[-1])
        pref_pdf_path = os.path.join(pref_dir, pref_pdf_files[-1]) if pref_pdf_files else None

        # Получаем файлы из NG
        ng_files = os.listdir(ng_dir)
        ng_excel_files = [f for f in ng_files if f.lower().endswith(".xlsx")]
        ng_pdf_files = [f for f in ng_files if f.lower().endswith(".pdf")]
        ng_excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(ng_dir, x)))
        ng_pdf_files.sort(key=lambda x: os.path.getmtime(os.path.join(ng_dir, x)))
        ng_excel_path = os.path.join(ng_dir, ng_excel_files[-1])
        ng_pdf_path = os.path.join(ng_dir, ng_pdf_files[-1]) if ng_pdf_files else None

        # Загружаем на сервер
        if pref_excel_path:
            upload_reports_to_server('Pref', [pref_excel_path] + ([pref_pdf_path] if pref_pdf_path else []))
        if ng_excel_path:
            upload_reports_to_server('NG', [ng_excel_path] + ([ng_pdf_path] if ng_pdf_path else []))

        print("Отчёты загружены на сервер")

        keep_latest_files(ng_dir, 'NG')
        keep_latest_files(pref_dir, 'Pref')
        clean_parcing_folder()
        print("Процесс завершен успешно!")
        _record_success('ng')

    except Exception as e:
        print(f"Ошибка в ng: {e}")
        import traceback
        traceback.print_exc()
        _record_failure('ng', str(e))
    finally:
        if _coinit:
            pythoncom.CoUninitialize()
        _running['ng'] = False
