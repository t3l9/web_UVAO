import os
import shutil
import time

import openpyxl
import pandas as pd
import pythoncom
import win32com.client
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService

from ..config import BASE_DIR, directory, login_NG, password_NG, _running
from ..utils.helpers import kill_excel_processes, upload_reports_to_server, keep_latest_files, clean_parcing_folder
from ..utils.status import _record_success, _record_failure, _get_chromedriver


def parcing_MWI(attempts=2):
    for attempt in range(1, attempts + 1):
        print(f"Попытка {attempt} из {attempts}")
        driver = webdriver.Chrome(service=ChromeService(_get_chromedriver()))
        processed_tabs = [False, False, False]
        try:
            driver.get('https://gorod.mos.ru/api/service/auth/auth')
            driver.maximize_window()
            username = driver.find_element(By.XPATH, '//input[@placeholder="Логин *"]')
            password = driver.find_element(By.XPATH, '//input[@placeholder="Пароль*"]')
            username.send_keys(login_NG)
            password.send_keys(password_NG)
            login_button = driver.find_element(By.XPATH,
                                               '/html/body/div[1]/div/div/main/div/div/div/div[2]/form[1]/button')
            login_button.click()
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                            '//div[@class="dashboard__block-link"]//div[@class="button-big link"]//div[@class="dashboard-container__links-title" and contains(text(), "Аналитика")]')))

            try:
                notification_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '/html/body/div/div/div[4]/div/div/div[2]/div[2]/div/div/button'))
                )
                notification_button.click()
                print("Уведомление закрыто")
                time.sleep(1)
            except Exception:
                print("Уведомление не обнаружено, продолжаем работу")

            time.sleep(0.5)
            driver.get('https://er.mos.ru/ker/admin/issues/monitor_mzi?sidebar=organization')
            time.sleep(4)

            try:
                button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//div[contains(@class, 'v-list-item__title') and text()='Ответы на доработке']"))
                )
                button.click()
                time.sleep(5)
                download_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[.//i[contains(@class, 'mdi-download')]]"))
                )
                download_btn.click()
                time.sleep(2)
                driver.execute_script("document.querySelector('.v-overlay__scrim.white').style.display='none';")
                time.sleep(1)
                driver.execute_script("document.querySelector('.v-overlay.v-overlay--active').style.display='none';")
                time.sleep(1)
                option = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '(//div[@class="v-select__selections"])[4]'))
                )
                option.click()
                time.sleep(2)
                option_to_select = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Все сообщения')]"))
                )
                option_to_select.click()
                time.sleep(2)
                export_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(span, 'Экспорт')]"))
                )
                export_button.click()
                time.sleep(3)
                processed_tabs[0] = True
                driver.get('https://er.mos.ru/ker/admin/issues/monitor_mzi?sidebar=organization')
                time.sleep(3)
            except Exception:
                print("❌ Первая вкладка не доступна для обработки. Пропускаем...")
                driver.get('https://er.mos.ru/ker/admin/issues/monitor_mzi?sidebar=organization')
                time.sleep(3)

            try:
                button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//div[contains(@class, 'v-list-item__title') and text()='Обещание устранения']"))
                )
                button.click()
                time.sleep(5)
                download_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[.//i[contains(@class, 'mdi-download')]]"))
                )
                download_btn.click()
                time.sleep(3)
                driver.execute_script("document.querySelector('.v-overlay__scrim.white').style.display='none';")
                time.sleep(1)
                driver.execute_script("document.querySelector('.v-overlay.v-overlay--active').style.display='none';")
                time.sleep(1)
                option = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '(//div[@class="v-select__selections"])[4]'))
                )
                option.click()
                time.sleep(2)
                option_to_select = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Все сообщения')]"))
                )
                option_to_select.click()
                time.sleep(2)
                export_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(span, 'Экспорт')]"))
                )
                export_button.click()
                time.sleep(3)
                processed_tabs[1] = True
                driver.get('https://er.mos.ru/ker/admin/issues/monitor_mzi?sidebar=organization')
                time.sleep(3)
            except Exception:
                print("❌ Вторая вкладка не доступна для обработки. Пропускаем...")
                driver.get('https://er.mos.ru/ker/admin/issues/monitor_mzi?sidebar=organization')
                time.sleep(3)

            try:
                button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(
                        (By.XPATH,
                         "//div[contains(@class, 'v-list-item__title') and text()='Нарушения для получателя']"))
                )
                button.click()
                time.sleep(3)
                download_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[.//i[contains(@class, 'mdi-download')]]"))
                )
                download_btn.click()
                time.sleep(3)
                driver.execute_script("document.querySelector('.v-overlay__scrim.white').style.display='none';")
                time.sleep(1)
                driver.execute_script("document.querySelector('.v-overlay.v-overlay--active').style.display='none';")
                time.sleep(1)
                option = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, '(//div[@class="v-select__selections"])[4]'))
                )
                option.click()
                time.sleep(2)
                option_to_select = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Все сообщения')]"))
                )
                option_to_select.click()
                time.sleep(2)
                export_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(span, 'Экспорт')]"))
                )
                export_button.click()
                time.sleep(3)
                processed_tabs[2] = True
            except Exception:
                print("❌ Третья вкладка не доступна для обработки. Пропускаем...")

            driver.get('https://gorod.mos.ru/admin/ker/olap/downloads')

            try:
                if processed_tabs[0]:
                    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH,
                                                                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[3]/td[5]/div/i')))
                    button = driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[3]/td[5]/div/i')
                    button.click()
                    time.sleep(1)
            except Exception:
                print("❌ Ошибка при скачивании файла для первой вкладки")
            try:
                if processed_tabs[1]:
                    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH,
                                                                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[2]/td[5]/div/i')))
                    button = driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[2]/td[5]/div/i')
                    button.click()
                    time.sleep(1)
            except Exception:
                print("❌ Ошибка при скачивании файла для второй вкладки")
            try:
                if processed_tabs[2]:
                    WebDriverWait(driver, 300).until(EC.presence_of_element_located((By.XPATH,
                                                                                     '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')))
                    button = driver.find_element(By.XPATH,
                                                 '/html/body/div[1]/div/div[2]/main/div/div[1]/div/div[2]/div[1]/table/tbody/tr[1]/td[5]/div/i')
                    button.click()
                    time.sleep(1.5)
            except Exception:
                print("❌ Ошибка при скачивании файла для третьей вкладки")

            return processed_tabs

        except Exception as e:
            print(f"❌ Произошла ошибка при выгрузке СВОДА МЖИ: {e}")
            driver.quit()
            if attempt == attempts:
                print("Парсинг не удался после всех попыток.")
                return
        finally:
            driver.quit()


def MWI_choosing_files(download_dir, processed_tabs_count):
    files = sorted([os.path.join(download_dir, f) for f in os.listdir(download_dir) if f.endswith('.xlsx')],
                   key=os.path.getmtime, reverse=True)

    if processed_tabs_count == 0:
        print("Нет файлов для обработки - ни одна вкладка не была выгружена")
        return None
    elif processed_tabs_count == 1:
        latest_files = files[:1]
    elif processed_tabs_count == 2:
        latest_files = files[:2]
    else:
        latest_files = files[:3]

    try:
        dfs = []
        for file in latest_files:
            if os.path.isfile(file):
                df = pd.read_excel(file)
                dfs.append(df)

        if not dfs:
            print("Не найдено ни одного файла для обработки")
            return None

        return pd.concat(dfs, ignore_index=True)

    except Exception as e:
        print(f"Ошибка при обработке файлов: {str(e)}")
        return None


def MWI_process_file(df):
    today = datetime.now()
    tomorrow = today + timedelta(days=1)
    day_after = today + timedelta(days=2)

    today_str = today.strftime('%d.%m.%Y')
    tomorrow_str = tomorrow.strftime('%d.%m.%Y')
    after_str = f"{day_after.strftime('%d.%m.%Y')} и далее"

    df['Тип'] = ''
    df['Дата отображения на мониторе'] = pd.to_datetime(df['Дата отображения на мониторе'], dayfirst=True,
                                                        format='%d.%m.%Y %H:%M:%S')
    df.loc[(df['Просрок Монитора'] == 'Да'), "Тип"] = "Просрок"
    df.loc[(df['Дата отображения на мониторе'].dt.date == today.date()) & (
            df['Просрок Монитора'] == 'Нет'), "Тип"] = today_str
    df.loc[(df['Дата отображения на мониторе'].dt.date == tomorrow.date()), "Тип"] = tomorrow_str
    df.loc[(df['Дата отображения на мониторе'].dt.date > tomorrow.date()), "Тип"] = after_str
    return df, today_str, tomorrow_str, after_str


def mwi():
    if _running['mwi']:
        print("[mwi] Пропуск: предыдущий запуск ещё выполняется")
        return
    _running['mwi'] = True
    _coinit = False
    try:
        today = datetime.now()
        timenow = today.strftime("%H-%M")

        print("🔍 [MWI] Начало парсинга...")
        raw_processed_tabs = parcing_MWI()
        processed_tabs = raw_processed_tabs if raw_processed_tabs else [False, False, False]
        processed_count = sum(processed_tabs)
        pythoncom.CoInitialize()
        _coinit = True

        print("🔍 [MWI] Обработка файлов...")
        df, today_str, tomorrow_str, after_str = MWI_process_file(MWI_choosing_files(directory, processed_count))

        print("🔍 [MWI] Проверка данных перед созданием Pivot...")

        if df is None or df.empty:
            print("❌ [MWI] Данные пустые! Пропускаем создание отчёта.")
            return

        print(f"✅ [MWI] Данные: {len(df)} строк, {len(df.columns)} столбцов")

        required_columns = ['Район', 'Тип', 'Номер заявки', 'Дата отображения на мониторе']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"❌ [MWI] Нет обязательных столбцов: {missing_columns}")
            return

        print("🔍 [MWI] Создание Excel файла...")
        filepath = os.path.join(directory, f"СВОД МЖИ {datetime.now().strftime('%d.%m.%Y')} {timenow}.xlsx")

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='МЖИ', index=False, startrow=0)

        if not os.path.exists(filepath):
            print(f"❌ [MWI] Файл не создан: {filepath}")
            return

        try:
            test_wb = openpyxl.load_workbook(filepath)
            test_ws = test_wb.active
            print(f"✅ [MWI] Файл корректен: {test_ws.max_row} строк, {test_ws.max_column} столбцов")
            test_wb.close()
        except Exception as e:
            print(f"❌ [MWI] Файл повреждён: {e}")
            return

        gen_date = datetime.now().strftime('%d.%m.%Y %H:%M')
        total_records = len(df)
        tab_count = processed_count
        tab1_icon = "✓" if processed_tabs[0] else "✗"
        tab2_icon = "✓" if processed_tabs[1] else "✗"
        tab3_icon = "✓" if processed_tabs[2] else "✗"

        print("🔍 [MWI] Создание Pivot Table (VBA)...")

        vba_macro = f"""
Sub CreatePivotTable()
    On Error GoTo MacroError
    Dim wsData As Worksheet
    Dim wsPivot As Worksheet
    Dim pivotCache As PivotCache
    Dim pivotTable As PivotTable
    Dim lastRow As Long
    Dim lastCol As Long

    Set wsData = ThisWorkbook.Sheets("МЖИ")

    With wsData.Columns("B")
        .NumberFormat = "DD.MM.YYYY"
    End With

    On Error Resume Next
    Application.DisplayAlerts = False
    ThisWorkbook.Sheets("Сводная таблица").Delete
    Application.DisplayAlerts = True
    On Error GoTo 0
    Set wsPivot = ThisWorkbook.Sheets.Add(Before:=ThisWorkbook.Sheets(1))
    wsPivot.Name = "Сводная таблица"

    lastCol = wsData.Cells(1, wsData.Columns.Count).End(xlToLeft).Column
    lastRow = wsData.Cells(wsData.Rows.Count, "A").End(xlUp).Row

    ' Подстраховка: если в столбце A есть дыры, меряем по самому длинному столбцу
    Dim c As Long, r As Long
    For c = 1 To lastCol
        r = wsData.Cells(wsData.Rows.Count, c).End(xlUp).Row
        If r > lastRow Then lastRow = r
    Next c

    ' Если данных нет (только шапка или пусто) — выходим без падения
    If lastRow < 2 Or lastCol < 1 Then
        Err.Raise vbObjectError + 513, "CreatePivotTable", "Нет данных для сводной: lastRow=" & lastRow & " lastCol=" & lastCol
    End If

    Dim srcAddr As String
    srcAddr = "'" & wsData.Name & "'!" & _
              wsData.Range(wsData.Cells(1, 1), wsData.Cells(lastRow, lastCol)).Address(True, True, xlR1C1)

    Set pivotCache = ThisWorkbook.PivotCaches.Create( _
        SourceType:=xlDatabase, _
        SourceData:=srcAddr, _
        Version:=6)

    Set pivotTable = pivotCache.CreatePivotTable( _
        TableDestination:=wsPivot.Cells(7, 1), _
        TableName:="МЖИСвод")

    With pivotTable
        .PivotFields("Район").Orientation = xlRowField
        .PivotFields("Тип").Orientation = xlColumnField
        .AddDataField .PivotFields("Номер заявки"), "Количество", xlCount
    End With

    wsPivot.Rows(7).Hidden = True
    wsPivot.Range("A8").Value = "Район"

    Dim typePivotField As PivotField
    Dim item As PivotItem
    Set typePivotField = pivotTable.PivotFields("Тип")
    For Each item In typePivotField.PivotItems
        If item.Name = "(blank)" Or item.Name = "" Then
            item.Visible = False
        End If
    Next item

    Dim pos As Integer
    pos = 1
    On Error Resume Next
    typePivotField.PivotItems("Просрок").Position = pos
    If Err.Number = 0 Then pos = pos + 1
    Err.Clear
    typePivotField.PivotItems("{today_str}").Position = pos
    If Err.Number = 0 Then pos = pos + 1
    Err.Clear
    typePivotField.PivotItems("{tomorrow_str}").Position = pos
    If Err.Number = 0 Then pos = pos + 1
    Err.Clear
    typePivotField.PivotItems("{after_str}").Position = pos
    Err.Clear
    On Error GoTo 0

    pivotTable.RefreshTable

    Dim lastPivotCol As Integer
    lastPivotCol = wsPivot.Cells(8, wsPivot.Columns.Count).End(xlToLeft).Column
    Dim totalRow As Long
    totalRow = wsPivot.Cells(wsPivot.Rows.Count, 1).End(xlUp).Row

    Dim rng As Range
    Set rng = wsPivot.Range(wsPivot.Cells(8, 1), wsPivot.Cells(totalRow, lastPivotCol))

    With rng
        .Font.Name = "Times New Roman"
        .Font.Size = 11
        .Interior.Color = RGB(255, 255, 255)
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
        .WrapText = True
    End With

    With rng.Borders
        .LineStyle = xlContinuous
        .Color = RGB(200, 205, 215)
        .Weight = xlThin
    End With

    rng.BorderAround xlContinuous, xlMedium, , RGB(100, 120, 170)

    Dim districtRange As Range
    Set districtRange = wsPivot.Range(wsPivot.Cells(9, 1), wsPivot.Cells(totalRow, 1))
    districtRange.HorizontalAlignment = xlLeft
    districtRange.IndentLevel = 1

    Dim colIdx As Integer
    For colIdx = 1 To lastPivotCol
        Dim hVal As String
        hVal = wsPivot.Cells(8, colIdx).Value
        Select Case hVal
            Case "Район"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(28, 54, 110)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                    .Font.Size = 12
                End With
            Case "Просрок"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(192, 40, 40)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                End With
            Case "{today_str}"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(200, 85, 15)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                End With
            Case "{tomorrow_str}"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(160, 120, 0)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                End With
            Case "{after_str}"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(35, 120, 55)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                End With
            Case "Общий итог"
                With wsPivot.Cells(8, colIdx)
                    .Interior.Color = RGB(65, 70, 100)
                    .Font.Color = RGB(255, 255, 255)
                    .Font.Bold = True
                End With
        End Select
    Next colIdx

    With wsPivot.Range(wsPivot.Cells(totalRow, 1), wsPivot.Cells(totalRow, lastPivotCol))
        .Font.Bold = True
        .Font.Size = 11
        .Interior.Color = RGB(215, 220, 240)
    End With
    wsPivot.Cells(totalRow, 1).HorizontalAlignment = xlLeft
    wsPivot.Cells(totalRow, 1).IndentLevel = 1

    Dim prosrokCol As Integer
    prosrokCol = -1
    For colIdx = 1 To lastPivotCol
        If wsPivot.Cells(8, colIdx).Value = "Просрок" Then
            prosrokCol = colIdx
            Exit For
        End If
    Next colIdx

    If prosrokCol > 0 Then
        Dim prosrokRng As Range
        Set prosrokRng = wsPivot.Range(wsPivot.Cells(9, prosrokCol), wsPivot.Cells(totalRow - 1, prosrokCol))
        prosrokRng.FormatConditions.Delete
        Dim fc As FormatCondition
        Set fc = prosrokRng.FormatConditions.Add(xlCellValue, xlGreater, 0)
        fc.Interior.Color = RGB(255, 195, 195)
        fc.Font.Color = RGB(140, 0, 0)
        fc.Font.Bold = True
    End If

    Dim rowIdx As Long
    For rowIdx = 9 To totalRow - 1
        If (rowIdx Mod 2) = 0 Then
            wsPivot.Range(wsPivot.Cells(rowIdx, 1), wsPivot.Cells(rowIdx, lastPivotCol)).Interior.Color = RGB(246, 248, 255)
        End If
    Next rowIdx

    wsPivot.Columns(1).ColumnWidth = 26
    For colIdx = 2 To lastPivotCol
        wsPivot.Columns(colIdx).ColumnWidth = 15
    Next colIdx

    wsPivot.Rows(8).RowHeight = 36
    For rowIdx = 9 To totalRow
        wsPivot.Rows(rowIdx).RowHeight = 22
    Next rowIdx

    wsPivot.Range("A1:H1").Merge
    wsPivot.Range("A2:H2").Merge
    wsPivot.Range("A3:H3").Merge

    With wsPivot.Cells(1, 1)
        .Value = "СВОД МЖИ  —  {gen_date}"
        .Font.Name = "Times New Roman"
        .Font.Size = 16
        .Font.Bold = True
        .Font.Color = RGB(20, 40, 100)
        .HorizontalAlignment = xlLeft
        .IndentLevel = 1
    End With
    wsPivot.Rows(1).RowHeight = 30

    With wsPivot.Cells(2, 1)
        .Value = "Всего записей: {total_records}   |   Загружено: {tab_count} из 3 вкладок"
        .Font.Name = "Times New Roman"
        .Font.Size = 11
        .Font.Color = RGB(80, 80, 90)
        .HorizontalAlignment = xlLeft
        .IndentLevel = 1
    End With
    wsPivot.Rows(2).RowHeight = 22

    With wsPivot.Cells(3, 1)
        .Value = "{tab1_icon}  Ответы на доработке       {tab2_icon}  Обещание устранения       {tab3_icon}  Нарушения для получателя"
        .Font.Name = "Times New Roman"
        .Font.Size = 11
        .Font.Color = RGB(60, 65, 80)
        .HorizontalAlignment = xlLeft
        .IndentLevel = 1
    End With
    wsPivot.Rows(3).RowHeight = 22

    With wsPivot.Range("A1:H3")
        .Interior.Color = RGB(238, 243, 255)
    End With

    With wsPivot.Range("A3:H3").Borders(xlEdgeBottom)
        .LineStyle = xlContinuous
        .Color = RGB(100, 130, 200)
        .Weight = xlMedium
    End With

    wsPivot.Rows(4).RowHeight = 5
    wsPivot.Rows(5).RowHeight = 5
    wsPivot.Rows(6).RowHeight = 5

    wsPivot.PageSetup.PrintArea = ""
    wsPivot.PageSetup.FitToPagesWide = 1
    wsPivot.PageSetup.FitToPagesTall = False
    wsPivot.PageSetup.Zoom = False
    wsPivot.PageSetup.LeftMargin = 14
    wsPivot.PageSetup.RightMargin = 14
    wsPivot.PageSetup.TopMargin = 14
    wsPivot.PageSetup.BottomMargin = 14

    Exit Sub
MacroError:
    Application.DisplayAlerts = False
    Err.Raise Err.Number, Err.Source, "MWI macro error: " & Err.Description
End Sub
"""

        print("🔍 [MWI] Запуск Excel...")
        excel = None
        workbook = None

        try:
            excel = win32com.client.Dispatch('Excel.Application')
            excel.Visible = False
            excel.DisplayAlerts = False

            print(f"🔍 [MWI] Открытие файла: {filepath}")
            workbook = excel.Workbooks.Open(filepath)

            sheet_names = [ws.Name for ws in workbook.Worksheets]
            if "МЖИ" not in sheet_names:
                print(f"❌ [MWI] Нет листа 'МЖИ'! Доступны: {sheet_names}")
                workbook.Close(SaveChanges=False)
                excel.Quit()
                return

            print("🔍 [MWI] Добавление VBA макроса...")
            vb_module = workbook.VBProject.VBComponents.Add(1)
            vb_module.CodeModule.AddFromString(vba_macro)

            print("🔍 [MWI] Выполнение макроса...")
            try:
                workbook.Application.Run("'CreatePivotTable'")
                print("✅ [MWI] Макрос выполнен")
            except Exception as macro_error:
                import traceback
                print(f"❌ [MWI] Ошибка макроса: {macro_error}")
                traceback.print_exc()
                workbook.Close(SaveChanges=False)
                excel.Quit()
                kill_excel_processes()
                return

            print("🔍 [MWI] Создание PDF...")
            pdf_file_name = os.path.join(directory, f"СВОД МЖИ {datetime.now().strftime('%d.%m.%Y')} {timenow}.pdf")
            pdf_path = os.path.join(os.path.dirname(filepath), pdf_file_name)

            wsFirst = workbook.Worksheets(1)
            wsFirst.PageSetup.FitToPagesWide = 1
            wsFirst.PageSetup.FitToPagesTall = 1
            wsFirst.PageSetup.Zoom = False
            wsFirst.PageSetup.LeftMargin = excel.Application.CentimetersToPoints(0.5)
            wsFirst.PageSetup.RightMargin = excel.Application.CentimetersToPoints(0.5)
            wsFirst.PageSetup.TopMargin = excel.Application.CentimetersToPoints(0.5)
            wsFirst.PageSetup.BottomMargin = excel.Application.CentimetersToPoints(0.5)

            workbook.Save()

            try:
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
                wsFirst.ExportAsFixedFormat(0, pdf_path)
                print(f"✅ [MWI] PDF создан: {pdf_path}")
            except Exception as e:
                print(f"⚠️ [MWI] Ошибка при создании PDF: {e}")
                pdf_path = None

            workbook.Close(SaveChanges=True)
            excel.Quit()
            print("✅ [MWI] Excel закрыт")

        except Exception as e:
            import traceback
            print(f"❌ [MWI] Ошибка Excel: {e}")
            traceback.print_exc()
            try:
                if workbook:
                    workbook.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
            except Exception:
                pass
            kill_excel_processes()
            raise

        print("🔍 [MWI] Перемещение файлов в public/MWI/...")
        static_directory = os.path.join(BASE_DIR, 'MWI')
        if not os.path.exists(static_directory):
            os.makedirs(static_directory)

        if os.path.exists(filepath):
            shutil.move(filepath, static_directory)
            print(f"✅ [MWI] Excel перемещен: {static_directory}")

        if pdf_path and os.path.exists(pdf_path):
            shutil.move(pdf_path, static_directory)
            print(f"✅ [MWI] PDF перемещен: {static_directory}")

        print("🔍 [MWI] Загрузка отчётов на сервер...")
        excel_path_final = os.path.join(static_directory, os.path.basename(filepath))
        pdf_path_final = os.path.join(static_directory, os.path.basename(pdf_path)) if pdf_path else None

        files_to_upload = [excel_path_final]
        if pdf_path_final and os.path.exists(pdf_path_final):
            files_to_upload.append(pdf_path_final)

        if upload_reports_to_server('MWI', files_to_upload):
            print("✔ [MWI] Отчёты MWI загружены на сервер")

        keep_latest_files(static_directory, 'MWI')
        clean_parcing_folder()
        print("✅ [MWI] Процесс завершен успешно!")
        _record_success('mwi')

    except Exception as e:
        import traceback
        print(f"❌ [MWI] Ошибка в mwi: {e}")
        traceback.print_exc()
        _record_failure('mwi', str(e))
    finally:
        if _coinit:
            pythoncom.CoUninitialize()
        _running['mwi'] = False
