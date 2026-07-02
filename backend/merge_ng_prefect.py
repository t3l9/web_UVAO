"""
Объединяет лист "Префект просрок" из всех исторических Excel в public/NG/
в один файл. Лист уже содержит только просрочки Префектуры ЮВАО.

Логика:
  1. Читает каждый "Ответы в работе DD.MM.YYYY HH-MM.xlsx", лист "Префект просрок"
  2. Добавляет столбец "Дата выгрузки" из имени файла
  3. Дедупликация по "Номер сообщения": оставляет строку из последнего файла
  4. Сохраняет итоговый Excel рядом со скриптом

Запуск:
    python merge_ng_prefect.py

Опции:
    --folder  путь к папке с Excel (по умолч.: backend/public/NG/)
    --output  путь к итоговому файлу (по умолч.: рядом со скриптом)
"""

import argparse
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FOLDER = os.path.join(SCRIPT_DIR, 'public', 'NG')
SHEET_NAME     = 'Префект просрок'

PREFERRED_COLS = [
    'Дата выгрузки',
    'Номер сообщения',
    'Номер заявки',
    'Дата публикации сообщения',
    'Регламентный срок у сообщения (Портал)',
    'Район',
    'Статус подготовки ответа на сообщение',
    'Адрес',
    'Проблемная тема',
    'Просрок (Монитор)',
    'Ответственный ОИВ первого уровня',
    'Ответственный за подготовку ответа',
]


def parse_export_date(filename: str):
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+(\d{2}-\d{2})', filename)
    if not m:
        return None, None
    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%d.%m.%Y %H-%M")
    return dt, dt.strftime('%d.%m.%Y %H:%M')


def run(ng_folder: str, output_path: str):
    if not os.path.isdir(ng_folder):
        print(f"Папка не найдена: {ng_folder}")
        return

    xlsx_files = [
        f for f in os.listdir(ng_folder)
        if f.lower().endswith('.xlsx') and f.startswith('Ответы в работе')
    ]

    parsed = []
    for fname in xlsx_files:
        dt, dt_str = parse_export_date(fname)
        if dt:
            parsed.append((dt, dt_str, os.path.join(ng_folder, fname)))
        else:
            print(f"  ? Пропущен (нет даты): {fname}")

    if not parsed:
        print("Нет файлов.")
        return

    parsed.sort(key=lambda x: x[0])
    print(f"Файлов: {len(parsed)}")
    print(f"Период: {parsed[0][0].strftime('%d.%m.%Y')} → {parsed[-1][0].strftime('%d.%m.%Y')}")
    print()

    frames = []

    for export_dt, export_str, filepath in parsed:
        fname = os.path.basename(filepath)
        print(f"  {fname} ...", end=' ', flush=True)

        try:
            df = pd.read_excel(filepath, sheet_name=SHEET_NAME)
        except Exception as e:
            print(f"нет листа / ошибка: {e}")
            continue

        if df.empty:
            print("пусто")
            continue

        df['_export_str'] = export_str
        df['_export_dt']  = export_dt
        print(f"{len(df)} строк")
        frames.append(df)

    if not frames:
        print(f"\nНет данных на листе «{SHEET_NAME}» ни в одном файле.")
        return

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nВсего строк до дедупликации: {len(combined)}")

    # Дедупликация: оставить последнее вхождение по номеру сообщения
    id_col = 'Номер сообщения' if 'Номер сообщения' in combined.columns else 'Номер заявки'
    if id_col in combined.columns:
        combined.sort_values('_export_dt', ascending=True, inplace=True)
        combined.drop_duplicates(subset=[id_col], keep='last', inplace=True)
    else:
        combined.drop_duplicates(inplace=True)

    print(f"Уникальных просрочек:        {len(combined)}")

    # Финальный вид
    combined.drop(columns=['_export_dt'], inplace=True, errors='ignore')
    combined.rename(columns={'_export_str': 'Дата выгрузки'}, inplace=True)

    # Форматируем дату дедлайна
    deadline_col = 'Регламентный срок у сообщения (Портал)'
    if deadline_col in combined.columns:
        combined[deadline_col] = pd.to_datetime(combined[deadline_col], errors='coerce') \
                                   .dt.strftime('%d.%m.%Y')

    # Упорядочиваем столбцы
    present = [c for c in PREFERRED_COLS if c in combined.columns]
    rest    = [c for c in combined.columns if c not in PREFERRED_COLS]
    combined = combined[present + rest]

    # Сортируем: самый старый дедлайн первым
    if deadline_col in combined.columns:
        combined['_sort'] = pd.to_datetime(combined[deadline_col], format='%d.%m.%Y', errors='coerce')
        combined.sort_values('_sort', ascending=True, inplace=True)
        combined.drop(columns=['_sort'], inplace=True)

    combined.to_excel(output_path, sheet_name='Префект просрок', index=False)

    # ── Оформление ──────────────────────────────────────────────────────────────
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    row_fill_odd = PatternFill('solid', fgColor='FFF0F0')   # лёгкий красноватый — просрочки Префекта
    row_fill_evn = PatternFill('solid', fgColor='FFFFFF')
    thin_side    = Side(style='thin', color='CBD5E1')
    border       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al      = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_al
        cell.border    = border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = row_fill_odd if row_idx % 2 else row_fill_evn
        for cell in row:
            cell.fill      = fill
            cell.alignment = left_al
            cell.border    = border

    col_widths = {
        'Дата выгрузки':                                 18,
        'Номер сообщения':                               20,
        'Номер заявки':                                  20,
        'Дата публикации сообщения':                     20,
        'Регламентный срок у сообщения (Портал)':        24,
        'Район':                                         18,
        'Статус подготовки ответа на сообщение':         30,
        'Адрес':                                         35,
        'Проблемная тема':                               35,
        'Просрок (Монитор)':                             20,
        'Ответственный ОИВ первого уровня':              30,
        'Ответственный за подготовку ответа':            30,
    }
    for col_idx, cell in enumerate(ws[1], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(str(cell.value), 20)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)

    size_kb = os.path.getsize(output_path) // 1024
    print(f"\nГотово!")
    print(f"  Строк: {len(combined)}")
    print(f"  Файл:  {output_path} ({size_kb} КБ)")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merge NG Prefect overdue sheet')
    parser.add_argument('--folder', default=DEFAULT_FOLDER)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()

    out = args.output or os.path.join(
        SCRIPT_DIR,
        f"Префект_просрочки_сводная_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    run(args.folder, out)
