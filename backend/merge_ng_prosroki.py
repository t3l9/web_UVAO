"""
Объединяет все исторические Excel из public/NG/ в один файл просрочек.

Логика:
  1. Читает каждый "Ответы в работе DD.MM.YYYY HH-MM.xlsx", лист "Ответы в работе"
  2. Добавляет столбец "Дата выгрузки" из имени файла
  3. Оставляет только строки, где Дата выгрузки > Регламентный срок (просрочены)
  4. Дедупликация по "Номер сообщения": оставляет последнее вхождение (самый свежий файл)
  5. Сохраняет итоговый Excel рядом со скриптом

Запуск (из папки backend/):
    python merge_ng_prosroki.py

Опции:
    --folder  путь к папке с Excel (по умолч.: backend/public/NG/)
    --output  путь к итоговому файлу (по умолч.: Просрочки_сводная_YYYYMMDD_HHMM.xlsx рядом со скриптом)
"""

import argparse
import os
import re
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FOLDER = os.path.join(SCRIPT_DIR, 'public', 'NG')

# Предпочтительный порядок столбцов в итоговом файле
PREFERRED_COLS = [
    'Дата выгрузки',
    'Номер сообщения',
    'Номер заявки',
    'Дата публикации сообщения',
    'Регламентный срок у сообщения (Портал)',
    'Район',
    'Статус подготовки ответа на сообщение',
    'Адрес',
    'Проблемная тема',
    'Просрок (Монитор)',
    'Ответственный ОИВ первого уровня',
    'Ответственный за подготовку ответа',
]


def parse_export_date(filename: str):
    """Возвращает (datetime, строку 'DD.MM.YYYY HH:MM') из имени файла."""
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})\s+(\d{2}-\d{2})', filename)
    if not m:
        return None, None
    dt = datetime.strptime(f"{m.group(1)} {m.group(2)}", "%d.%m.%Y %H-%M")
    return dt, dt.strftime('%d.%m.%Y %H:%M')


def load_file(filepath: str, export_dt: datetime, export_str: str) -> pd.DataFrame | None:
    """Читает лист 'Ответы в работе', добавляет столбец Дата выгрузки."""
    try:
        df = pd.read_excel(filepath, sheet_name='Ответы в работе')
    except Exception as e:
        print(f"  ! Не удалось прочитать лист: {e}")
        return None

    if df.empty:
        return None

    df['Дата выгрузки'] = export_dt          # datetime для сравнений
    df['_export_str']   = export_str          # строка для отображения
    df['_export_dt']    = export_dt           # для сортировки/дедупликации
    return df


def run(ng_folder: str, output_path: str):
    if not os.path.isdir(ng_folder):
        print(f"Папка не найдена: {ng_folder}")
        return

    xlsx_files = [
        f for f in os.listdir(ng_folder)
        if f.lower().endswith('.xlsx') and f.startswith('Ответы в работе')
    ]

    parsed = []
    for fname in xlsx_files:
        dt, dt_str = parse_export_date(fname)
        if dt:
            parsed.append((dt, dt_str, os.path.join(ng_folder, fname)))
        else:
            print(f"  ? Пропущен (нет даты в имени): {fname}")

    if not parsed:
        print("Нет файлов для обработки.")
        return

    parsed.sort(key=lambda x: x[0])  # от старых к новым
    print(f"Найдено файлов: {len(parsed)}")
    print(f"Период: {parsed[0][0].strftime('%d.%m.%Y')} → {parsed[-1][0].strftime('%d.%m.%Y')}")
    print()

    frames = []
    for export_dt, export_str, filepath in parsed:
        fname = os.path.basename(filepath)
        print(f"  Читаю: {fname} ...", end=' ', flush=True)
        df = load_file(filepath, export_dt, export_str)
        if df is None:
            print("пусто")
            continue

        # Определяем столбец с регламентным сроком
        deadline_col = 'Регламентный срок у сообщения (Портал)'
        if deadline_col not in df.columns:
            print(f"нет столбца '{deadline_col}', пропуск")
            continue

        # Приводим срок к datetime
        df[deadline_col] = pd.to_datetime(df[deadline_col], errors='coerce')

        # Оставляем только просрочки: дата выгрузки > регламентный срок
        mask = df['_export_dt'] > df[deadline_col]
        overdue = df[mask].copy()

        print(f"строк всего {len(df)}, просрочек {len(overdue)}")
        if not overdue.empty:
            frames.append(overdue)

    if not frames:
        print("\nНет просрочек ни в одном файле.")
        return

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nВсего строк до дедупликации: {len(combined)}")

    # Определяем ключевой столбец для дедупликации
    id_col = 'Номер сообщения' if 'Номер сообщения' in combined.columns else 'Номер заявки'

    if id_col not in combined.columns:
        print(f"Столбец '{id_col}' не найден — дедупликация по полному совпадению строк")
        combined.drop_duplicates(inplace=True)
    else:
        # Сортируем по дате выгрузки (возрастание) — keep='last' оставит самое свежее
        combined.sort_values('_export_dt', ascending=True, inplace=True)
        combined.drop_duplicates(subset=[id_col], keep='last', inplace=True)

    print(f"Уникальных просрочек:       {len(combined)}")

    # Формируем итоговый DataFrame
    # Убираем служебные столбцы
    combined.drop(columns=['Дата выгрузки', '_export_dt'], inplace=True, errors='ignore')
    combined.rename(columns={'_export_str': 'Дата выгрузки'}, inplace=True)

    # Форматируем регламентный срок как дату
    deadline_col = 'Регламентный срок у сообщения (Портал)'
    if deadline_col in combined.columns:
        combined[deadline_col] = pd.to_datetime(combined[deadline_col], errors='coerce') \
                                   .dt.strftime('%d.%m.%Y')

    # Упорядочиваем столбцы
    present_preferred = [c for c in PREFERRED_COLS if c in combined.columns]
    rest = [c for c in combined.columns if c not in PREFERRED_COLS]
    combined = combined[present_preferred + rest]

    # Сортируем итог: сначала самый просроченный (самый ранний дедлайн)
    if deadline_col in combined.columns:
        combined[deadline_col] = pd.to_datetime(combined[deadline_col], format='%d.%m.%Y', errors='coerce')
        combined.sort_values(deadline_col, ascending=True, inplace=True)
        combined[deadline_col] = combined[deadline_col].dt.strftime('%d.%m.%Y')

    # ── Сохранение ──────────────────────────────────────────────────────────────
    combined.to_excel(output_path, sheet_name='Просрочки', index=False)

    # ── Оформление ──────────────────────────────────────────────────────────────
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    row_fill_odd = PatternFill('solid', fgColor='EEF2FF')
    row_fill_evn = PatternFill('solid', fgColor='FFFFFF')
    thin_side    = Side(style='thin', color='CBD5E1')
    border       = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Заголовок
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = border

    # Строки данных
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = row_fill_odd if row_idx % 2 else row_fill_evn
        for cell in row:
            cell.fill      = fill
            cell.alignment = left_align
            cell.border    = border

    # Ширина столбцов
    col_widths = {
        'Дата выгрузки':                                 18,
        'Номер сообщения':                               20,
        'Номер заявки':                                  20,
        'Дата публикации сообщения':                     20,
        'Регламентный срок у сообщения (Портал)':        24,
        'Район':                                         18,
        'Статус подготовки ответа на сообщение':         30,
        'Адрес':                                         35,
        'Проблемная тема':                               35,
        'Просрок (Монитор)':                             20,
        'Ответственный ОИВ первого уровня':              30,
        'Ответственный за подготовку ответа':            30,
    }

    for col_idx, cell in enumerate(ws[1], start=1):
        letter = get_column_letter(col_idx)
        width  = col_widths.get(str(cell.value), 20)
        ws.column_dimensions[letter].width = width

    # Заморозить шапку
    ws.freeze_panes = 'A2'

    # Авто-фильтр
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)

    size_kb = os.path.getsize(output_path) // 1024
    print(f"\nГотово!")
    print(f"  Строк в итоге: {len(combined)}")
    print(f"  Файл: {output_path} ({size_kb} КБ)")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merge overdue NG reports into one Excel')
    parser.add_argument(
        '--folder', default=DEFAULT_FOLDER,
        help=f'Папка с файлами (по умолч.: {DEFAULT_FOLDER})'
    )
    parser.add_argument(
        '--output', default=None,
        help='Путь к итоговому файлу (по умолч.: рядом со скриптом)'
    )
    args = parser.parse_args()

    out = args.output or os.path.join(
        SCRIPT_DIR,
        f"Просрочки_сводная_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    run(args.folder, out)
